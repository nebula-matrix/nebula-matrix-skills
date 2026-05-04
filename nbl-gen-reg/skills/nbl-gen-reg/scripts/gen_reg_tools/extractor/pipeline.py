import os
import re
from typing import List, Optional, Tuple

import xlrd

from gen_reg_tools.core.data_model import Block, Top
from gen_reg_tools.extractor.excel_reader import ExcelReader, _SheetAdapter
from gen_reg_tools.extractor.cif_mapper import CifAddrMapper


class Extractor:
    """寄存器数据提取器"""

    DEFAULT_PREFIX = "ASIC_PDT_V3R1"

    @staticmethod
    def _build_regex(prefix: Optional[str] = None) -> re.Pattern:
        """构建文件名匹配正则，不区分大小写"""
        if prefix is None:
            prefix = Extractor.DEFAULT_PREFIX
        escaped = re.escape(prefix)
        pattern = rf"^{escaped}_([^_]+)_((?:[^_]+_?)+)\.xlsx$"
        return re.compile(pattern, re.IGNORECASE)

    @staticmethod
    def _parse_filename(file_name: str, prefix: Optional[str] = None) -> Tuple[str, str]:
        """解析文件名，返回 (it_name, bt_name)。不匹配时返回空字符串"""
        regex = Extractor._build_regex(prefix)
        match = regex.match(file_name)
        if match:
            return match.group(1), match.group(2)
        return "", ""

    @staticmethod
    def _check_filename_format(file_name: str, prefix: Optional[str] = None) -> None:
        """检查以指定前缀开头但格式错误的文件，抛出 ValueError"""
        if prefix is None:
            prefix = Extractor.DEFAULT_PREFIX

        # 忽略非 xlsx 文件
        if not file_name.lower().endswith(".xlsx"):
            return

        prefix_lower = prefix.lower()
        name_lower = os.path.splitext(file_name)[0].lower()

        if not name_lower.startswith(prefix_lower + "_"):
            return

        remaining = name_lower[len(prefix_lower) + 1:]
        parts = remaining.split("_")

        if len(parts) < 2 or any(not p.strip() for p in parts[:2]):
            raise ValueError(
                f"文件名格式错误: '{file_name}' 以 '{prefix}' 开头，"
                f"但未包含 it_name 和 bt_name 两部分。期望格式: "
                f"{prefix}_{{it_name}}_{{bt_name}}.xlsx"
            )

    @staticmethod
    def _is_valid_sheet(sheet) -> bool:
        """检查 sheet 是否包含寄存器表头"""
        sheet_row_num = sheet.nrows if sheet.nrows <= 5 else 5
        for line in range(sheet_row_num):
            line_txt = sheet.row_values(line)
            if ("offset_addr" in line_txt) and ("name" in line_txt) and ("type" in line_txt):
                return True
        return False

    @staticmethod
    def _extract_file_blocks(file_path: str, cif_path: str, it_name: str) -> List[Block]:
        """从单个 Excel 文件提取所有有效的 sheet，返回 Block 列表"""
        # 优先 xlrd，回退 openpyxl；均失败则跳过该文件
        try:
            wb = xlrd.open_workbook(file_path)
            backend = "xlrd"
            sheet_ls = wb.sheet_names()
        except Exception:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                backend = "openpyxl"
                sheet_ls = wb.sheetnames
            except Exception:
                return []

        blocks: List[Block] = []

        for mdl in sheet_ls:
            if backend == "xlrd":
                raw_sheet = wb.sheet_by_name(mdl)
            else:
                raw_sheet = wb[mdl]

            sheet = _SheetAdapter(raw_sheet, backend)
            if not Extractor._is_valid_sheet(sheet):
                continue

            reader = ExcelReader(file_path, mdl)
            registers = reader.parse()

            block = Block(
                name=mdl,
                offset_addr=None,
                block_type="bt",
                subpart=registers,
            )
            blocks.append(block)

        return blocks

    @staticmethod
    def extract(path: str, cif_path: str, prefix: Optional[str] = None) -> Top:
        """统一入口，自动判断文件或目录"""
        if os.path.isfile(path):
            return Extractor.extract_file(path, cif_path, prefix)
        elif os.path.isdir(path):
            return Extractor.extract_dir(path, cif_path, prefix)
        else:
            raise FileNotFoundError(f"Path not found: {path}")

    @staticmethod
    def extract_file(excel_path: str, cif_path: str, prefix: Optional[str] = None) -> Top:
        """从单个 Excel 文件提取"""
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        file_name = os.path.basename(excel_path)

        # 检查格式错误（以 prefix_ 开头但缺少 it/bt）
        Extractor._check_filename_format(file_name, prefix)

        it_name, bt_name = Extractor._parse_filename(file_name, prefix)
        if not it_name:
            it_name = "IT_NAME"

        blocks = Extractor._extract_file_blocks(excel_path, cif_path, it_name)
        return Top(name="TOP", subpart=blocks)

    @staticmethod
    def extract_dir(excel_dir: str, cif_path: str, prefix: Optional[str] = None) -> Top:
        """从目录批量提取"""
        if not os.path.isdir(excel_dir):
            raise NotADirectoryError(f"Not a directory: {excel_dir}")

        all_blocks: List[Block] = []

        for root, dirs, files in os.walk(excel_dir):
            for file_name in files:
                # 先检查格式错误（以 prefix_ 开头但缺少 it/bt）
                Extractor._check_filename_format(file_name, prefix)

                it_name, bt_name = Extractor._parse_filename(file_name, prefix)
                if it_name:
                    file_abs = os.path.join(excel_dir, file_name)
                    blocks = Extractor._extract_file_blocks(file_abs, cif_path, it_name)
                    all_blocks.extend(blocks)

        return Top(name="TOP", subpart=all_blocks)
