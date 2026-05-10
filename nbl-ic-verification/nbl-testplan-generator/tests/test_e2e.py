"""End-to-end test: spec_tree.json -> combine_writer.py -> xlsx."""
import json
from pathlib import Path

import pytest

from parsers.md_parser import parse_spec_tree
from writers.combine_writer import write_testplan_xlsx


class TestE2E:
    def test_md_parser_to_spec_tree_json(self, temp_dir, sample_spec_md):
        """md_parser should produce valid spec_tree.json."""
        result = parse_spec_tree(sample_spec_md)
        assert result["module_name"] == "upa"
        assert "chapters" in result
        assert len(result["chapters"]) > 0

        # Save as JSON for next stage
        json_path = temp_dir / "spec_tree.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        assert json_path.exists()

    def test_spec_tree_to_xlsx(self, temp_dir, sample_template_xlsx):
        """Full pipeline: spec_tree-like JSON -> xlsx."""
        testplan_data = {
            "module_name": "upa",
            "spec_name": "upa_spec",
            "chapter": "chapter 1 功能特性",
            "review_items": [
                {
                    "item_id": "R1",
                    "feature": "数据通路特性",
                    "question": "是否支持小于64B的数据包？",
                    "source": "spec ch3 PA.001",
                    "status": "pending",
                }
            ],
            "features": [
                {
                    "feature": "数据通路特性",
                    "feature_id": "PA.001",
                    "_eng_id": "dp",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "单向数据传输",
                            "_eng_id": "unidir",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2": "正常场景",
                                    "_confidence": "confirmed",
                                    "subfeatures_l3": [
                                        {
                                            "subfeature_l3": "数据从源端到目的端",
                                            "remarks": "来源: PA.001 单向数据传输；相关寄存器: upa_ctrl.en",
                                            "stimulus": "【配置】upa_ctrl.en=1\n【激励】发送标准数据包",
                                            "checking": "by_checker",
                                            "coverage": "cp_upa_en: {0,1}",
                                            "priority": "HIGH",
                                            "activity": "BT",
                                            "path": "",
                                        }
                                    ],
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output = temp_dir / "e2e_output.xlsx"
        result = write_testplan_xlsx(testplan_data, sample_template_xlsx, output)
        assert output.exists()

        # Verify readable xlsx
        import openpyxl
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws.max_row > 1

        # Verify B column chapter marker
        found_chapter = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if val and "chapter 1" in str(val):
                found_chapter = True
                break
        assert found_chapter, "Chapter marker not found"

        # Verify feature in D column
        found_feature = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=4).value
            if val and "数据通路特性" in str(val):
                found_feature = True
                break
        assert found_feature, "Feature not found"

        # Verify W column path uses _eng_id
        found_path = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=23).value
            if val and "cg_dp" in str(val):
                found_path = True
                break
        assert found_path, "W column _eng_id path not found"
