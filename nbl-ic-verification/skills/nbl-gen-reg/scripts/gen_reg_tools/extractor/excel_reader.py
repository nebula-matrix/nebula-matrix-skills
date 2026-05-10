import re
from collections import defaultdict
from typing import Any, Dict, List, Tuple

import xlrd

from gen_reg_tools.core.data_model import Field, Register
from gen_reg_tools.core.utils import RegUtils


class _SheetAdapter:
    """统一 xlrd 和 openpyxl sheet 接口"""

    def __init__(self, sheet, backend: str):
        self._sheet = sheet
        self._backend = backend
        if backend == "openpyxl":
            self.nrows = sheet.max_row
            self.ncols = sheet.max_column
        else:
            self.nrows = sheet.nrows
            self.ncols = sheet.ncols

    def row_values(self, row: int) -> List[Any]:
        if self._backend == "openpyxl":
            if row >= self.nrows:
                return []
            return [
                self._sheet.cell(row=row + 1, column=col + 1).value or ""
                for col in range(self.ncols)
            ]
        return self._sheet.row_values(row)

    def cell_value(self, row: int, col: int) -> Any:
        if self._backend == "openpyxl":
            return self._sheet.cell(row=row + 1, column=col + 1).value or ""
        return self._sheet.cell_value(row, col)

    @property
    def merged_cells(self):
        if self._backend == "openpyxl":
            # openpyxl: ranges 是 CellRange 对象，属性为 1-based
            for mr in self._sheet.merged_cells.ranges:
                yield (mr.min_row - 1, mr.max_row - 1, mr.min_col - 1, mr.max_col - 1)
        else:
            # xlrd: 已经是 0-based (start_row, end_row, start_col, end_col)
            for item in self._sheet.merged_cells:
                yield item


class ExcelReader:
    """Excel 寄存器表解析引擎（支持 xlrd 和 openpyxl）"""

    def __init__(self, path: str, sheet_name: str) -> None:
        self.path = path
        self.sheet_name = sheet_name
        self.multick_ls: List[str] = []
        self.all_cells_cnt = 0

        # 优先尝试 xlrd（.xls），失败回退到 openpyxl（.xlsx）
        try:
            wb = xlrd.open_workbook(path)
            sheet = wb.sheet_by_name(sheet_name)
            self._backend = "xlrd"
        except Exception:
            from openpyxl import load_workbook

            wb = load_workbook(path, data_only=True)
            sheet = wb[sheet_name]
            self._backend = "openpyxl"

        self.wb = wb
        self.sheet = _SheetAdapter(sheet, self._backend)

    def get_merge_cells(self) -> Dict[Tuple[int, int], Any]:
        """获取合并单元格映射，键为 (row, col)，值为合并单元格的内容"""
        merge_cell_dict: Dict[Tuple[int, int], Any] = {}
        for (start_row, end_row, start_col, end_col) in self.sheet.merged_cells:
            value = self.sheet.cell_value(start_row, start_col)
            for row in range(start_row, end_row):
                for col in range(start_col, end_col):
                    merge_cell_dict[(row, col)] = value
        return merge_cell_dict

    @staticmethod
    def _get_msb_lsb(bits_range: str) -> Tuple[int, int]:
        """解析 bit 范围字符串，返回 (lsb, bits_num)
        例如 "15:0" -> (0, 16)，"7" -> (7, 1)
        """
        num_ls = re.findall(r"\d+", bits_range)
        num_ls = [int(item) for item in num_ls]
        if len(num_ls) == 1:
            msb_lsb = 1
        else:
            msb_lsb = num_ls[0] - num_ls[1] + 1
        return (num_ls[-1], msb_lsb)

    def extract_data(self):
        """逐行提取 Excel 数据，生成 (clock_domain, reg_dict, field_dict)"""
        previous_row_res = []
        start_extract_line = 1
        extract_sheet_flag = False

        # 定位表头行（前 5 行内搜索）
        sheet_row_num = self.sheet.nrows if self.sheet.nrows <= 5 else 5
        for line in range(sheet_row_num):
            line_txt = self.sheet.row_values(line)
            if ("offset_addr" in line_txt) and ("name" in line_txt) and ("type" in line_txt):
                extract_sheet_flag = True
                break
            else:
                start_extract_line += 1

        if not extract_sheet_flag:
            raise ValueError(
                f"Sheet '{self.sheet_name}' cannot be extracted: missing required headers"
            )

        title_ls = self.sheet.row_values(start_extract_line - 1)
        exist_multick_domain = "clock_domain" in title_ls

        merge_cell_dict = self.get_merge_cells()
        self.all_cells_cnt = len(title_ls) * self.sheet.nrows

        for row in range(start_extract_line, self.sheet.nrows):
            row_ls = self.sheet.row_values(row)

            # 跳过空行
            if not RegUtils.check_list(row_ls[:8]):
                continue

            row_res_ls = []
            for title in title_ls:
                data_org = merge_cell_dict.get(
                    (row, title_ls.index(title)), row_ls[title_ls.index(title)]
                )
                if title == "offset_addr" or title == "default_value":
                    data = RegUtils.addr_to_int(data_org)
                    row_res_ls.append(data)
                else:
                    if isinstance(data_org, str):
                        data = data_org.lower().strip()
                    else:
                        data = int(data_org)
                    row_res_ls.append(data)

            # 处理时钟域
            if exist_multick_domain:
                clock_domain = merge_cell_dict.get(
                    (row, title_ls.index("clock_domain")),
                    row_ls[title_ls.index("clock_domain")],
                )
                clock_domain = clock_domain.replace("clk", "")
                self.multick_ls.append(clock_domain)
            else:
                clock_domain = ""

            # 前 5 列（offset_addr, name, type, depth, width）空值向前继承
            for idx in range(5):
                row_res_ls[idx] = (
                    row_res_ls[idx] if row_res_ls[idx] != "" else previous_row_res[idx]
                )

            previous_row_res = row_res_ls

            reg_dict_keys = title_ls[:5]
            reg_dict_vals = row_res_ls[:5]

            field_dict_keys = title_ls[5:]
            field_dict_vals = row_res_ls[5:]

            # snapshot_en 处理：若存在则归入寄存器属性，否则默认 na
            if "snapshot_en" in field_dict_keys:
                snapshot_idx = field_dict_keys.index("snapshot_en")
                reg_dict_keys.append(field_dict_keys[snapshot_idx])
                reg_dict_vals.append(field_dict_vals[snapshot_idx])
            else:
                reg_dict_keys.append("snapshot_en")
                reg_dict_vals.append("na")

            # 过滤掉 offset_addr 和 name 均为空的行
            if "" not in reg_dict_vals[:2]:
                yield (
                    clock_domain,
                    dict(zip(reg_dict_keys, reg_dict_vals)),
                    dict(zip(field_dict_keys, field_dict_vals)),
                )

    @staticmethod
    def _rm_repeat_dicts(dict_list: List[Dict]) -> List[Dict]:
        """去除列表中重复的字典（基于键值对排序后的元组）"""
        seen = set()
        unique_list = []
        for d in dict_list:
            sorted_items = tuple(sorted(d.items()))
            if sorted_items not in seen:
                seen.add(sorted_items)
                unique_list.append(d)
        return unique_list

    def parse(self) -> List[Register]:
        """解析整张 sheet，返回 Register 对象列表"""
        merge_dict: Dict[Tuple[int, str], List[Dict]] = defaultdict(list)
        remain_reg_dict: Dict[Tuple[int, str, str], Dict] = {}

        # 第一遍：收集所有行数据
        for clock_domain, reg_line_dict, field_line_dict in self.extract_data():
            lsb, msb_lsb = self._get_msb_lsb(field_line_dict["range"])

            key = (reg_line_dict["offset_addr"], reg_line_dict["name"].lower().strip())

            # 构建 field 信息字典
            field_info = dict(field_line_dict)
            field_info["bits"] = int(msb_lsb)
            field_info["lsb"] = int(lsb)
            field_info["type"] = "field"
            field_info["level"] = "field"
            field_info["name"] = field_line_dict["field"]

            # 构建寄存器剩余属性字典
            remain = {
                "type": reg_line_dict["type"].lower().strip(),
                "snapshot_en": reg_line_dict["snapshot_en"].lower().strip(),
                "depth": RegUtils.addr_to_int(reg_line_dict["depth"]),
                "width": RegUtils.addr_to_int(reg_line_dict["width"]),
                "clock_domain": clock_domain.lower().strip(),
            }

            merge_dict[key].append(field_info)

            if (
                remain["type"] != ""
                and remain["depth"] != ""
                and remain["width"] != ""
            ):
                remain_key = (
                    reg_line_dict["offset_addr"],
                    reg_line_dict["name"].lower().strip(),
                    "remain_reg",
                )
                remain_reg_dict[remain_key] = remain

        # 第二遍：组装 Register / Field 对象
        registers: List[Register] = []
        for (offset_addr, name), field_items in merge_dict.items():
            remain = remain_reg_dict.get((offset_addr, name, "remain_reg"), {})

            # 去重 field（去除重复的 rsv 等）
            unique_fields = self._rm_repeat_dicts(field_items)

            # 计算 vld_bits 和 tbl_rw_attr
            vld_bits = 0
            tbl_rw_attr = None
            reg_type = remain.get("type", "")
            for f in unique_fields:
                if reg_type in ["tbl", "wide_reg"]:
                    vld_bits += f["bits"] if f.get("field") != "rsv" else 0
                    if reg_type == "tbl":
                        tbl_rw_attr = (
                            "rw"
                            if f.get("rw_attr") == "rw" or tbl_rw_attr == "rw"
                            else None
                        )

            # 构建 Field 对象
            field_objects: List[Field] = []
            for f in unique_fields:
                default_val = f.get("default_value", 0)
                if default_val == "" or default_val is None:
                    default_val = 0
                elif isinstance(default_val, str):
                    converted = RegUtils.addr_to_int(default_val)
                    default_val = converted if isinstance(converted, int) else 0

                field_objects.append(
                    Field(
                        name=f.get("name", ""),
                        bits=f.get("bits", 0),
                        lsb=f.get("lsb", 0),
                        rw_attr=f.get("rw_attr", ""),
                        default_value=default_val,
                        wr_ctrl=f.get("wr_ctrl"),
                        fcov=f.get("fcov"),
                    )
                )

            reg_vld_bits = vld_bits if reg_type in ["tbl", "wide_reg"] else None

            registers.append(
                Register(
                    name=name,
                    offset_addr=offset_addr,
                    width=remain.get("width", 0),
                    depth=remain.get("depth", 0),
                    reg_type=reg_type,
                    clock_domain=remain.get("clock_domain", ""),
                    vld_bits=reg_vld_bits,
                    snapshot_en=remain.get("snapshot_en", "na"),
                    subpart=field_objects,
                )
            )

        return registers
