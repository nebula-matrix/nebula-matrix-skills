#!/usr/bin/env python3
"""reg_parser.py - Parse register xlsx into reg_info.json structure."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any


def parse_register_xlsx(filepath: str | Path) -> dict[str, Any]:
    """Parse a register xlsx into structured reg_info dict."""
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Register file not found: {filepath}")

    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    module_name = _infer_module_name(filepath.stem)

    result: dict[str, Any] = {
        "module_name": module_name,
        "source_file": filepath.name,
        "registers": [],
        "register_tables": [],
    }

    skip_sheets = {"封面", "地址映射", "address_map", "summary", "sheet1"}

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in {s.lower() for s in skip_sheets}:
            continue
        ws = wb[sheet_name]
        registers = _parse_register_sheet(ws)
        result["registers"].extend(registers)

    wb.close()
    return result


def _parse_register_sheet(ws) -> list[dict[str, Any]]:
    """Parse a single worksheet containing register definitions."""
    registers: list[dict[str, Any]] = []

    # Detect header row
    header_row_num = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        row_values = [str(v).lower().strip() if v else "" for v in row]
        header_keywords = ["offset", "addr", "address", "name", "field", "bit", "description"]
        if sum(1 for kv in header_keywords if any(kv in rv for rv in row_values)) >= 2:
            header_row_num = i
            break

    if header_row_num is None:
        return registers

    # Build column index from header
    header = [str(v).strip() if v else "" for v in ws.iter_rows(
        min_row=header_row_num, max_row=header_row_num, values_only=True
    ).__next__()]

    col_map: dict[str, int] = {}
    for idx, h in enumerate(header):
        hl = h.lower()
        if "offset" in hl or "addr" in hl:
            col_map["addr"] = idx
        elif "name" in hl and "field" not in hl:
            col_map["reg_name"] = idx
        elif "field" in hl:
            col_map["field_name"] = idx
        elif "bit" in hl:
            col_map["bit_range"] = idx
        elif "rw" in hl or "access" in hl:
            col_map["rw"] = idx
        elif "description" in hl or "desc" in hl:
            col_map["desc"] = idx
        elif "reset" in hl or "default" in hl:
            col_map["reset"] = idx

    current_reg: dict[str, Any] | None = None

    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        row_values = [str(v) if v is not None else "" for v in row]
        if not any(row_values):
            continue

        reg_name = row_values[col_map.get("reg_name", 0)].strip() if "reg_name" in col_map else ""
        field_name = row_values[col_map.get("field_name", 0)].strip() if "field_name" in col_map else ""

        if reg_name and (not field_name or field_name.lower() == reg_name.lower()):
            if current_reg is not None:
                registers.append(current_reg)
            addr = row_values[col_map.get("addr", 0)] if "addr" in col_map else ""
            desc = row_values[col_map.get("desc", -1)] if "desc" in col_map else ""
            current_reg = {
                "reg_name": reg_name,
                "reg_addr": addr,
                "reg_desc": desc,
                "fields": [],
            }
            if field_name and field_name.lower() != reg_name.lower():
                _add_field(current_reg, row_values, col_map)
        elif current_reg is not None and field_name:
            _add_field(current_reg, row_values, col_map)

    if current_reg is not None:
        registers.append(current_reg)

    return registers


def _add_field(reg: dict, row_values: list, col_map: dict) -> None:
    """Append a field to the current register."""
    field = {
        "field_name": row_values[col_map.get("field_name", 0)] if "field_name" in col_map else "",
        "bit_range": row_values[col_map.get("bit_range", 0)] if "bit_range" in col_map else "",
        "rw": row_values[col_map.get("rw", 0)] if "rw" in col_map else "",
        "reset": row_values[col_map.get("reset", 0)] if "reset" in col_map else "",
        "desc": row_values[col_map.get("desc", -1)] if "desc" in col_map else "",
    }
    reg["fields"].append(field)


def _infer_module_name(stem: str) -> str:
    """Infer module name from filename."""
    stem_lower = stem.lower()
    for name in ("upa", "dped", "uvn"):
        if name in stem_lower:
            return name
    m = re.search(r"[a-z]+", stem_lower)
    return m.group(0) if m else "module"


def get_register_by_name(reg_info: dict, reg_name: str) -> dict | None:
    """Find a register by name in reg_info."""
    for reg in reg_info.get("registers", []):
        if reg.get("reg_name") == reg_name:
            return reg
    return None


def get_field_by_name(reg: dict, field_name: str) -> dict | None:
    """Find a field by name within a register."""
    for field in reg.get("fields", []):
        if field.get("field_name") == field_name:
            return field
    return None
