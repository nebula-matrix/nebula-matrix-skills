#!/usr/bin/env python3
"""combine_writer.py - Write testplan JSON data into xlsx template.

Handles D-E-F-G hierarchy, fonts (微软雅黑), borders, red markers for 【配置】/【激励】,
and _eng_id-based path generation for W column.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any


def write_testplan_xlsx(
    testplan_data: dict[str, Any],
    template_path: str | Path,
    output_path: str | Path,
    verify_level: str = "BT",
) -> Path:
    """Write testplan JSON data into an xlsx template.

    Parameters
    ----------
    testplan_data : dict
        Must contain: module_name, spec_name, chapter, features[]
        Each feature: feature, feature_id, _eng_id (optional), subfeatures_l1[]
        Each subfeatures_l1: subfeature_l1, _eng_id (optional), subfeatures_l2[]
        Each subfeatures_l2: subfeature_l2, _confidence, subfeatures_l3[]
        Each subfeatures_l3: subfeature_l3, remarks, stimulus, checking,
                             coverage, priority, activity, path
    template_path : str | Path
        Path to testplan_template.xlsx
    output_path : str | Path
        Output xlsx file path
    verify_level : str
        UT / BT / IT / ST / EMU (default BT)

    Returns
    -------
    Path
        Path to the generated output file
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, Side

    STD_FONT = Font(name="微软雅黑", size=10)
    RED_FONT = Font(name="微软雅黑", size=10, color="FF0000")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    LEFT_ALIGN_VCENTER = Alignment(vertical="center", horizontal="left", wrap_text=True)

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # B4: spec name
    ws.cell(row=4, column=2).value = f"{testplan_data.get('module_name', 'module')}_spec"

    data_start_row = _find_data_start_row(ws)
    content_min_row = data_start_row
    content_max_row = data_start_row
    current_row = data_start_row

    # Write B column chapter marker once
    ws.cell(row=current_row, column=2).value = "chapter 1 功能特性"
    _apply_style(ws.cell(row=current_row, column=2), STD_FONT, LEFT_ALIGN_VCENTER)
    current_row += 1

    features = testplan_data.get("features", [])
    module_name = testplan_data.get("module_name", "module")

    for feature in features:
        # D column: Feature (outline level 0)
        ws.cell(row=current_row, column=4).value = feature.get("feature", "")
        _apply_style(ws.cell(row=current_row, column=4), STD_FONT, LEFT_ALIGN_VCENTER)
        current_row += 1

        feature_eng_id = feature.get("_eng_id", "")

        for sub_l1 in feature.get("subfeatures_l1", []):
            # E column: Subfeature L1 (outline level 1)
            ws.cell(row=current_row, column=5).value = sub_l1.get("subfeature_l1", "")
            _apply_style(ws.cell(row=current_row, column=5), STD_FONT, LEFT_ALIGN_VCENTER)
            current_row += 1

            sub_l1_eng_id = sub_l1.get("_eng_id", feature_eng_id)

            for sub_l2 in sub_l1.get("subfeatures_l2", []):
                # F column: Subfeature L2 (outline level 2)
                ws.cell(row=current_row, column=6).value = sub_l2.get("subfeature_l2", "")
                _apply_style(ws.cell(row=current_row, column=6), STD_FONT, LEFT_ALIGN_VCENTER)

                confidence = sub_l2.get("_confidence", "confirmed")
                l3_list = sub_l2.get("subfeatures_l3", [])
                sub_l2_eng_id = sub_l2.get("_eng_id", sub_l1_eng_id)

                if l3_list:
                    current_row += 1
                    for sub_l3 in l3_list:
                        ws.cell(row=current_row, column=7).value = sub_l3.get("subfeature_l3", "")
                        row_data = dict(sub_l3)
                        row_data.setdefault("_feature_eng_id", feature_eng_id)
                        row_data.setdefault("_subfeature_eng_id", sub_l2_eng_id)
                        row_data["_module"] = module_name
                        _write_data_row(ws, current_row, row_data, verify_level, confidence, STD_FONT, RED_FONT, LEFT_ALIGN_VCENTER)
                        current_row += 1
                else:
                    row_data = dict(sub_l2)
                    row_data["subfeature_l3"] = ""
                    row_data.setdefault("_feature_eng_id", feature_eng_id)
                    row_data.setdefault("_subfeature_eng_id", sub_l2_eng_id)
                    row_data["_module"] = module_name
                    _write_data_row(ws, current_row, row_data, verify_level, confidence, STD_FONT, RED_FONT, LEFT_ALIGN_VCENTER)
                    current_row += 1

        content_max_row = max(content_max_row, current_row)

    _apply_borders(ws, content_min_row, content_max_row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path.resolve()


def _find_data_start_row(ws) -> int:
    for row in range(6, 50):
        val = ws.cell(row=row, column=4).value
        if val is None or str(val).strip() == "":
            return row
    return 6


def _write_data_row(ws, row, data, verify_level, confidence, std_font, red_font, alignment):
    is_inferred = confidence == "inferred"
    font = red_font if is_inferred else std_font

    # H: remarks
    ws.cell(row=row, column=8).value = data.get("remarks", "")
    ws.cell(row=row, column=8).font = font
    ws.cell(row=row, column=8).alignment = alignment

    # I: stimulus
    stimulus = data.get("stimulus", "")
    ws.cell(row=row, column=9).value = stimulus
    ws.cell(row=row, column=9).alignment = alignment
    _color_stimulus_markers(ws, row, 9, stimulus)

    # J: checking
    ws.cell(row=row, column=10).value = data.get("checking", "by_checker")
    ws.cell(row=row, column=10).font = font
    ws.cell(row=row, column=10).alignment = alignment

    # K: coverage
    ws.cell(row=row, column=11).value = data.get("coverage", "")
    ws.cell(row=row, column=11).font = font
    ws.cell(row=row, column=11).alignment = alignment

    # M: activity
    ws.cell(row=row, column=13).value = data.get("activity", verify_level)
    ws.cell(row=row, column=13).font = font
    ws.cell(row=row, column=13).alignment = alignment

    # W: path
    path = data.get("path", "")
    if not path:
        feature_eng_id = data.get("_feature_eng_id", "")
        subfeature_eng_id = data.get("_subfeature_eng_id", "")
        module = data.get("_module", "module")
        checking = data.get("checking", "by_checker")
        path = _build_path(feature_eng_id, subfeature_eng_id, checking, module)
    ws.cell(row=row, column=23).value = path
    ws.cell(row=row, column=23).font = font
    ws.cell(row=row, column=23).alignment = alignment


def _color_stimulus_markers(ws, row, col, text):
    from openpyxl.styles import Font
    if "【配置】" in text or "【激励】" in text:
        ws.cell(row=row, column=col).font = Font(name="微软雅黑", size=10, color="FF0000")
    else:
        ws.cell(row=row, column=col).font = Font(name="微软雅黑", size=10)


def _build_path(feature_eng_id, subfeature_eng_id, checking, module="module"):
    if not feature_eng_id:
        return ""
    sub = subfeature_eng_id or feature_eng_id
    if checking == "by_checker":
        return f"Group:$unit::{module}_fcov::cg_{feature_eng_id}.cp_{sub}"
    elif checking == "by_direct_tc":
        return f"Group:$unit::{module}_direct_fcov::direct_{feature_eng_id}_{sub}"
    elif checking == "by_assertion":
        return f"Group:$unit::{module}_assert::assert_{feature_eng_id}_{sub}"
    return ""


def _apply_style(cell, font, alignment):
    cell.font = font
    cell.alignment = alignment


def _apply_borders(ws, min_row, max_row):
    from openpyxl.styles import Border, Side
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in range(min_row, max_row + 1):
        for col in range(1, 24):
            ws.cell(row=row, column=col).border = border


if __name__ == "__main__":
    import json
    import sys
    if len(sys.argv) < 4:
        print("Usage: python combine_writer.py <testplan.json> <template.xlsx> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)
    result = write_testplan_xlsx(data, sys.argv[2], sys.argv[3])
    print(f"Written: {result}")
