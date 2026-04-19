#!/usr/bin/env python3
"""cfg_writer.py - Configuration feature generation for Ch2.

This module handles Ch2 (配置特性) generation with two modes:
1. Independent mode: Generate Ch2 only from register JSON
2. Incremental mode: Append Ch2 to existing func_xlsx

Features:
- Register filtering (exclude *_int_*, *_fifo_*, etc.)
- Register grouping (_0/_1/_2 -> _0/1/2)
- Skip marking for covered registers
- Cross-reference with Ch1 features
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side

# Standard font: Microsoft YaHei 10pt
STD_FONT = Font(name="\u5fae\u8f6f\u96c5\u9ed1", size=10)  # 微软雅黑
STD_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column indices
COL_D = 4  # Register name (outline level 0)
COL_E = 5  # Subfeature L1 (outline level 1)
COL_F = 6  # Subfeature L2 overview (outline level 2)
COL_G = 7  # Subfeature L2 detail (outline level 3)
COL_H = 8  # Remarks
COL_I = 9  # Stimulus
COL_J = 10  # Checking
COL_K = 11  # Coverage
COL_L = 12  # Priority
COL_M = 13  # Activity
COL_W = 23  # Path

# Exclusion patterns for Ch2 registers
CFG_EXCLUDE_PATTERNS: list[str] = [
    r".*_int_status$",
    r".*_int_mask$",
    r".*_int_set$",
    r".*_int_.*$",
    r".*_car_ctrl$",
    r".*_fifo_th$",
    r".*_fifo_.*$",
    r".*_spare\d*$",
    r".*_cnt$",
    r".*_fsm\d*$",
    r".*_err_info$",
    r".*_bp_state$",
    r".*_bp_history$",
    r".*_test$",
    r".*_init_done$",
    r".*_init_start$",
]

_CFG_EXCLUDE_COMPILED = [re.compile(p) for p in CFG_EXCLUDE_PATTERNS]

# Valid access types for configuration registers
VALID_CFG_ACCESS = {"RW", "RWC", "WO", "RWW"}


def _is_excluded_register(reg_name: str) -> bool:
    """Check if register should be excluded from Ch2."""
    for pat in _CFG_EXCLUDE_COMPILED:
        if pat.match(reg_name):
            return True
    return False


def filter_cfg_registers(registers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Filter registers for Ch2 configuration features.

    - Excludes registers matching exclusion patterns
    - Keeps only fields with valid access types (RW, RWC, WO, RWW)

    Parameters
    ----------
    registers : list[dict]
        List of register dictionaries from reg_to_json output.

    Returns
    -------
    list[dict]
        Filtered registers with valid configuration fields.
    """
    result: list[dict[str, Any]] = []

    for reg in registers:
        reg_name = reg.get("name", "")

        # Skip excluded registers
        if _is_excluded_register(reg_name):
            continue

        # Filter fields to only valid access types
        valid_fields = []
        for field in reg.get("fields", []):
            access = field.get("access", "").upper()
            if access in VALID_CFG_ACCESS:
                valid_fields.append(field)

        # Only include register if it has valid fields
        if valid_fields:
            result.append({
                **reg,
                "fields": valid_fields,
            })

    return result


def group_similar_registers(registers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Group registers that differ only by a trailing numeric suffix.

    E.g. upa_stage_0, upa_stage_1, upa_stage_2 -> upa_stage_0/1/2

    Parameters
    ----------
    registers : list[dict]
        List of register dictionaries.

    Returns
    -------
    list[dict]
        Grouped registers with merged names.
    """
    groups: dict[str, list[dict[str, Any]]] = {}
    order: list[str] = []

    for reg in registers:
        name = reg["name"]
        m = re.match(r"^(.+?)_(\d+)$", name)
        if m:
            base = m.group(1)
        else:
            base = name

        if base not in groups:
            groups[base] = []
            order.append(base)
        groups[base].append(reg)

    result: list[dict[str, Any]] = []
    for base in order:
        members = groups[base]
        if len(members) == 1:
            result.append(members[0])
        else:
            suffixes = []
            for m_reg in members:
                m_match = re.match(r"^.+?_(\d+)$", m_reg["name"])
                suffixes.append(m_match.group(1) if m_match else "")

            merged_name = f"{base}_{'/'.join(suffixes)}"

            # Merge fields from all members
            all_fields: list[dict[str, Any]] = []
            seen: set[str] = set()
            for m_reg in members:
                for field in m_reg.get("fields", []):
                    key = field.get("name", "")
                    if key and key not in seen:
                        seen.add(key)
                        all_fields.append(field)

            result.append({
                "name": merged_name,
                "offset": members[0].get("offset", ""),
                "fields": all_fields,
                "description": members[0].get("description", ""),
            })

    return result


def extract_covered_regs_from_xlsx(func_path: str) -> set[str]:
    """Extract covered register names from existing func xlsx.

    Scans the stimulus and coverage columns for register references.

    Parameters
    ----------
    func_path : str
        Path to the func xlsx file.

    Returns
    -------
    set[str]
        Set of register names found in the xlsx.
    """
    wb = openpyxl.load_workbook(func_path, data_only=True)
    ws = wb.active

    covered: set[str] = set()

    # Find header row by looking for "stimulus" in headers
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for j, cell in enumerate(row):
            if cell and "stimulus" in str(cell).lower():
                header_row = i
                break
        if header_row:
            break

    if header_row is None:
        wb.close()
        return covered

    # Find column indices for stimulus and coverage
    headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    col_map = {}
    for j, h in enumerate(headers):
        if h:
            h_lower = str(h).lower()
            if "stimulus" in h_lower:
                col_map["stimulus"] = j + 1
            elif "coverage" in h_lower:
                col_map["coverage"] = j + 1

    # Default columns if not found
    stimulus_col = col_map.get("stimulus", COL_I)
    coverage_col = col_map.get("coverage", COL_K)

    # Scan data rows for register names (upa_xxx pattern)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        stimulus = str(row[stimulus_col - 1] or "") if len(row) >= stimulus_col else ""
        coverage = str(row[coverage_col - 1] or "") if len(row) >= coverage_col else ""
        text = stimulus + " " + coverage

        # Find register names (upa_xxx pattern)
        for match in re.finditer(r"(upa_\w+)", text):
            reg_name = match.group(1)
            # Normalize: strip trailing _0/_1 if present
            base_match = re.match(r"^(.+?)_\d+$", reg_name)
            if base_match:
                covered.add(base_match.group(1))
            covered.add(reg_name)

    wb.close()
    return covered


def cross_ref_cfg(
    ch2_regs: list[dict[str, Any]],
    covered_regs: set[str],
) -> list[dict[str, Any]]:
    """Cross-reference Ch2 registers with covered registers.

    Parameters
    ----------
    ch2_regs : list[dict]
        Ch2 register list (after filtering and grouping).
    covered_regs : set[str]
        Set of already covered register names.

    Returns
    -------
    list[dict]
        Cross-referenced registers with skip flag and stimulus.
    """
    result: list[dict[str, Any]] = []

    for reg in ch2_regs:
        reg_name = reg["name"]
        fields = reg.get("fields", [])

        # Check if this register (or its base) is covered
        is_covered = reg_name in covered_regs
        if not is_covered:
            # Check if grouped name matches
            base_match = re.match(r"^(.+?)_\d+(/\d+)*$", reg_name)
            if base_match:
                base = base_match.group(1)
                if base in covered_regs:
                    is_covered = True
                # Also check individual members
                for covered in covered_regs:
                    if covered.startswith(base + "_"):
                        is_covered = True
                        break

        if is_covered:
            # Mark as skip
            result.append({
                "name": reg_name,
                "skip": True,
                "stimulus": "",
                "checking": "",
                "coverage": "",
                "path": "",
                "remarks": "已在功能特性中覆盖",
            })
        else:
            # Generate stimulus
            desc_parts = []
            config_parts = []
            for field in fields:
                fname = field.get("name", "")
                fdesc = field.get("description", "")
                frange = field.get("range", "")
                faccess = field.get("access", "")

                if fname and fname.lower() != "rsv":
                    if fdesc:
                        desc_parts.append(f"{frange} {fname}: {fdesc}")

                    # Generate config for RW/RWC/WO fields
                    if faccess.upper() in VALID_CFG_ACCESS:
                        config_parts.append(f"配置 {reg_name}.{fname} ({frange})")

            stimulus = "【配置】" + "；".join(config_parts) if config_parts else ""

            result.append({
                "name": reg_name,
                "skip": False,
                "stimulus": stimulus,
                "checking": "",
                "coverage": "",
                "path": "【反标路径暂无法确定】",
                "remarks": "; ".join(desc_parts) if desc_parts else "",
            })

    return result


def build_ch2_json(
    reg_data: dict[str, Any],
    covered_regs: set[str] | None = None,
) -> dict[str, Any]:
    """Build Ch2 JSON data structure.

    Parameters
    ----------
    reg_data : dict
        Register JSON data from reg_to_json output.
    covered_regs : set[str] | None
        Set of already covered register names (optional).

    Returns
    -------
    dict
        Ch2 data structure ready for xlsx_writer.
    """
    if covered_regs is None:
        covered_regs = set()

    registers = reg_data.get("registers", [])
    filtered = filter_cfg_registers(registers)
    grouped = group_similar_registers(filtered)
    xrefed = cross_ref_cfg(grouped, covered_regs)

    ch2_registers: list[dict[str, Any]] = []
    for reg_info in xrefed:
        reg_name = reg_info["name"]
        is_skip = reg_info.get("skip", False)

        subfeatures_l2: list[dict[str, Any]] = [{
            "subfeature_l2_overview": "",  # F column: empty for Ch2
            "subfeature_l2_detail": "",    # G column: empty for Ch2
            "remarks": reg_info.get("remarks", ""),
            "stimulus": reg_info.get("stimulus", ""),
            "checking": reg_info.get("checking", ""),
            "coverage": reg_info.get("coverage", ""),
            "priority": "HIGH" if not is_skip else "",
            "activity": "BT",
            "path": reg_info.get("path", ""),
            "skip": is_skip,
        }]

        ch2_registers.append({
            "register_name": reg_name,
            "subfeatures_l1": [{
                "subfeature_l1": reg_name,
                "subfeatures_l2": subfeatures_l2,
            }],
        })

    return {
        "chapter": "chapter 2 配置特性",
        "registers": ch2_registers,
    }


def write_independent_ch2_xlsx(
    reg_data: dict[str, Any],
    output_path: str,
    template_path: str,
) -> dict[str, int]:
    """Write independent Ch2 xlsx (no func_xlsx reference).

    Parameters
    ----------
    reg_data : dict
        Register JSON data from reg_to_json output.
    output_path : str
        Destination xlsx file path.
    template_path : str
        Path to the template xlsx.

    Returns
    -------
    dict[str, int]
        Statistics: total_registers, covered_registers, new_registers.
    """
    # Import from same directory as this script
    import importlib.util
    _xlsx_writer_path = Path(__file__).parent / "xlsx_writer.py"
    _spec = importlib.util.spec_from_file_location("xlsx_writer_local", _xlsx_writer_path)
    _xlsx_writer = importlib.util.module_from_spec(_spec)
    _spec.loader.exec_module(_xlsx_writer)
    write_ch1_xlsx = _xlsx_writer.write_ch1_xlsx

    ch2_data = build_ch2_json(reg_data, covered_regs=set())

    # Use xlsx_writer to write Ch2 data (treating it as Ch1 structure)
    # Note: xlsx_writer.write_ch1_xlsx works with the same hierarchy structure
    write_ch1_xlsx(ch2_data, template_path, output_path)

    return {
        "total_registers": len(ch2_data.get("registers", [])),
        "covered_registers": 0,
        "new_registers": len(ch2_data.get("registers", [])),
    }


def append_cfg_to_func_xlsx(
    func_path: str,
    reg_data: dict[str, Any],
    output_path: str,
) -> dict[str, int]:
    """Append Ch2 configuration features to existing func xlsx.

    Parameters
    ----------
    func_path : str
        Path to the existing func xlsx.
    reg_data : dict
        Register JSON data from reg_to_json output.
    output_path : str
        Path for the output xlsx.

    Returns
    -------
    dict[str, int]
        Statistics: total_registers, covered_registers, new_registers.
    """
    # Load func xlsx
    wb = openpyxl.load_workbook(func_path)
    ws = wb.active

    # Extract covered registers from func xlsx
    covered_regs = extract_covered_regs_from_xlsx(func_path)

    # Build Ch2 data with cross-reference
    ch2_data = build_ch2_json(reg_data, covered_regs)

    # Find the last data row (row with data in D column)
    last_row = ws.max_row
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=COL_D).value:
            last_row = row
            break

    # Append Ch2 data
    next_row = last_row + 1

    # Add Ch2 chapter header row
    ws.cell(row=next_row, column=COL_D, value="chapter 2 配置特性")
    ws.row_dimensions[next_row].outline_level = 0
    next_row += 1

    covered_count = 0
    new_count = 0

    for reg in ch2_data.get("registers", []):
        reg_name = reg.get("register_name", "")

        for sf1 in reg.get("subfeatures_l1", []):
            for sf2 in sf1.get("subfeatures_l2", []):
                is_skip = sf2.get("skip", False)

                # Register row (D column)
                ws.cell(row=next_row, column=COL_D, value=reg_name)
                ws.cell(row=next_row, column=COL_D).font = STD_FONT
                ws.row_dimensions[next_row].outline_level = 1
                next_row += 1

                # Data row (E column empty, F-M columns filled)
                ws.cell(row=next_row, column=COL_H, value=sf2.get("remarks", ""))
                ws.cell(row=next_row, column=COL_H).font = STD_FONT
                ws.cell(row=next_row, column=COL_I, value=sf2.get("stimulus", ""))
                ws.cell(row=next_row, column=COL_I).font = STD_FONT
                ws.cell(row=next_row, column=COL_J, value=sf2.get("checking", ""))
                ws.cell(row=next_row, column=COL_J).font = STD_FONT
                ws.cell(row=next_row, column=COL_K, value=sf2.get("coverage", ""))
                ws.cell(row=next_row, column=COL_K).font = STD_FONT
                ws.cell(row=next_row, column=COL_L, value=sf2.get("priority", ""))
                ws.cell(row=next_row, column=COL_L).font = STD_FONT
                ws.cell(row=next_row, column=COL_M, value=sf2.get("activity", ""))
                ws.cell(row=next_row, column=COL_M).font = STD_FONT
                ws.cell(row=next_row, column=COL_W, value=sf2.get("path", ""))
                ws.cell(row=next_row, column=COL_W).font = STD_FONT

                # Apply borders
                for col in range(COL_D, COL_W + 1):
                    ws.cell(row=next_row, column=col).border = STD_BORDER

                ws.row_dimensions[next_row].outline_level = 2
                next_row += 1

                if is_skip:
                    covered_count += 1
                else:
                    new_count += 1

    # Save output
    wb.save(output_path)
    wb.close()

    return {
        "total_registers": covered_count + new_count,
        "covered_registers": covered_count,
        "new_registers": new_count,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate Ch2 configuration features",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Independent mode (Ch2 only)
  python cfg_writer.py --reg reg.json --output ch2.xlsx --template template.xlsx

  # Incremental mode (append to func_xlsx)
  python cfg_writer.py --reg reg.json --output combined.xlsx --func-xlsx func.xlsx
        """,
    )
    parser.add_argument("--reg", required=True, help="Path to register JSON")
    parser.add_argument("--output", required=True, help="Output xlsx path")
    parser.add_argument("--template", help="Template xlsx path (for independent mode)")
    parser.add_argument(
        "--func-xlsx",
        help="Path to existing func xlsx (for incremental mode)",
    )
    args = parser.parse_args()

    # Load register data
    with open(args.reg, "r", encoding="utf-8") as f:
        reg_data = json.load(f)

    # Determine mode
    if args.func_xlsx:
        # Incremental mode
        stats = append_cfg_to_func_xlsx(args.func_xlsx, reg_data, args.output)
        print(f"Incremental mode: {stats}")
    elif args.template:
        # Independent mode
        stats = write_independent_ch2_xlsx(reg_data, args.output, args.template)
        print(f"Independent mode: {stats}")
    else:
        parser.error("Either --template (independent mode) or --func-xlsx (incremental mode) is required")

    print(f"Output written to: {args.output}")
