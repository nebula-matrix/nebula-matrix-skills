#!/usr/bin/env python3
"""xlsx_writer.py - Write testplan features to Excel output.

Creates xlsx files with D-E-F-G hierarchy using outline levels 0-3.
H-W columns are only written at minimum granularity (G if exists, else F).
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter

# Standard font: Microsoft YaHei 10pt
STD_FONT = Font(name="\u5fae\u8f6f\u96c5\u9ed1", size=10)  # 微软雅黑
STD_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column indices
COL_D = 4  # Feature (outline level 0)
COL_E = 5  # Subfeature L1 (outline level 1)
COL_F = 6  # Subfeature L2 (outline level 2)
COL_G = 7  # Subfeature L3 (outline level 3)
COL_H = 8  # Remarks/Details
COL_I = 9  # Stimulus
COL_J = 10  # Checking
COL_K = 11  # Coverage
COL_L = 12  # Priority
COL_M = 13  # Activity
COL_W = 23  # Path to Source

# Data columns: H through W
DATA_COLS = list(range(COL_H, COL_W + 1))


def write_ch1_xlsx(
    features_data: dict[str, Any],
    template_path: Path | str,
    output_path: Path | str,
) -> Path:
    """Write Chapter 1 features to an Excel file.

    Parameters
    ----------
    features_data : dict
        Features data structure with module_name, features list, etc.
    template_path : Path | str
        Path to template xlsx file.
    output_path : Path | str
        Path for output xlsx file.

    Returns
    -------
    Path
        Path to the created output file.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Load template
    wb = load_workbook(template_path)
    ws = wb.active

    # Find the starting row (after template header rows)
    # Template has rows 1-5 as header, row 6 starts chapter section
    # We'll find the first empty row after row 6
    start_row = _find_data_start_row(ws)

    # Write each feature
    features = features_data.get("features", [])
    for feature in features:
        start_row = _write_feature(ws, feature, start_row)

    # Save output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return output_path


def _find_data_start_row(ws) -> int:
    """Find the starting row for data entry (after template headers)."""
    # Skip header rows (rows 1-5 in template)
    # Start checking from row 7 where first data typically goes
    for row in range(7, 100):  # Check up to row 100
        cell_b = ws.cell(row=row, column=2)  # Column B (Section #)
        if cell_b.value is None:
            # Check if this is truly empty (no data in D-W)
            has_data = False
            for col in range(COL_D, COL_W + 1):
                if ws.cell(row=row, column=col).value:
                    has_data = True
                    break
            if not has_data:
                return row
    return 7


def _write_feature(ws, feature: dict[str, Any], row: int) -> int:
    """Write a feature and its subfeatures to the worksheet.

    Returns the next available row.
    """
    feature_name = feature.get("feature", "")
    feature_id = feature.get("feature_id", "")

    # Check if feature has subfeatures_l3 (G level)
    has_l3 = _has_level_3(feature)

    # Write feature row (D column)
    cell_d = ws.cell(row=row, column=COL_D)
    display_name = f"{feature_name}\n{feature_id}" if feature_id else feature_name
    cell_d.value = display_name
    cell_d.font = STD_FONT
    ws.row_dimensions[row].outline_level = 0

    # Apply borders to D column
    for col in range(COL_D, COL_G + 1):  # D-G for this row
        ws.cell(row=row, column=col).border = STD_BORDER

    # Move to next row for subfeatures
    row += 1

    # Write L1 subfeatures
    for subfeature_l1 in feature.get("subfeatures_l1", []):
        row = _write_subfeature_l1(ws, subfeature_l1, row, has_l3)

    return row


def _has_level_3(feature: dict[str, Any]) -> bool:
    """Check if feature has level 3 subfeatures."""
    for l1 in feature.get("subfeatures_l1", []):
        for l2 in l1.get("subfeatures_l2", []):
            if l2.get("subfeatures_l3"):
                return True
    return False


def _write_subfeature_l1(ws, subfeature: dict[str, Any], row: int, has_l3: bool) -> int:
    """Write L1 subfeature and its children."""
    subfeature_name = subfeature.get("subfeature_l1", "")

    cell_e = ws.cell(row=row, column=COL_E)
    cell_e.value = subfeature_name
    cell_e.font = STD_FONT
    ws.row_dimensions[row].outline_level = 1

    # Apply borders
    for col in range(COL_D, COL_G + 1):
        ws.cell(row=row, column=col).border = STD_BORDER

    row += 1

    # Write L2 subfeatures
    for l2 in subfeature.get("subfeatures_l2", []):
        row = _write_subfeature_l2(ws, l2, row, has_l3)

    return row


def _write_subfeature_l2(ws, subfeature: dict[str, Any], row: int, has_l3: bool) -> int:
    """Write L2 subfeature and its children."""
    overview = subfeature.get("subfeature_l2_overview", "")
    detail = subfeature.get("subfeature_l2_detail", "")
    l3_list = subfeature.get("subfeatures_l3", [])

    # Build display value for F column
    if overview and detail:
        display_value = f"{overview}\n{detail}"
    else:
        display_value = overview or detail or ""

    cell_f = ws.cell(row=row, column=COL_F)
    cell_f.value = display_value
    cell_f.font = STD_FONT
    ws.row_dimensions[row].outline_level = 2

    # Apply borders
    for col in range(COL_D, COL_G + 1):
        ws.cell(row=row, column=col).border = STD_BORDER

    # If no L3, this is minimum granularity - write H-W columns
    if not l3_list:
        _write_data_columns(ws, subfeature, row)
        row += 1
    else:
        # Has L3, write L3 children
        row += 1
        for l3 in l3_list:
            row = _write_subfeature_l3(ws, l3, row)

    return row


def _write_subfeature_l3(ws, subfeature: dict[str, Any], row: int) -> int:
    """Write L3 subfeature (minimum granularity)."""
    subfeature_name = subfeature.get("subfeature_l3", "")

    cell_g = ws.cell(row=row, column=COL_G)
    cell_g.value = subfeature_name
    cell_g.font = STD_FONT
    ws.row_dimensions[row].outline_level = 3

    # Apply borders
    for col in range(COL_D, COL_W + 1):
        ws.cell(row=row, column=col).border = STD_BORDER

    # This is minimum granularity - write H-W columns
    _write_data_columns(ws, subfeature, row)

    return row + 1


def _write_data_columns(ws, data: dict[str, Any], row: int) -> None:
    """Write data columns H-W for minimum granularity rows."""
    # Column mappings
    column_mapping = {
        COL_H: data.get("remarks") or data.get("subfeature_l2_detail", ""),
        COL_I: data.get("stimulus", ""),
        COL_J: data.get("checking", ""),
        COL_K: data.get("coverage", ""),
        COL_L: data.get("priority", ""),
        COL_M: data.get("activity", ""),
        COL_W: data.get("path", ""),
    }

    for col, value in column_mapping.items():
        if value:
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.font = STD_FONT
            cell.border = STD_BORDER


if __name__ == "__main__":
    import json
    import sys

    if len(sys.argv) < 4:
        print("Usage: python xlsx_writer.py <features.json> <template.xlsx> <output.xlsx>")
        sys.exit(1)

    features_path = Path(sys.argv[1])
    template_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    with open(features_path, "r", encoding="utf-8") as f:
        features_data = json.load(f)

    result = write_ch1_xlsx(features_data, template_path, output_path)
    print(f"Output written to: {result}")
