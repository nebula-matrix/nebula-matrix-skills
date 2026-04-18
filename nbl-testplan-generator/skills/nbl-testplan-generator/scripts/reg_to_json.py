#!/usr/bin/env python3
"""reg_to_json.py - Convert register xlsx to structured JSON.

This is the refactored version of reg_parser.py with simplified output format.
"""
import json
import sys
import warnings
from pathlib import Path
from typing import Any

import openpyxl


def parse_register_sheet(ws) -> list[dict[str, Any]]:
    """Parse a single register sheet, returning a list of register dicts."""
    registers = []
    header_row_num = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row[0] and "offset_addr" in str(row[0]).lower():
            header_row_num = i
            break
        if len(row) > 1 and row[1] and "name" in str(row[1]).lower():
            header_row_num = i
            break

    if header_row_num is None:
        warnings.warn(
            f"Header row not found in sheet '{ws.title}' within first 5 rows. "
            "Assuming data starts from row 3.",
            UserWarning,
        )
        header_row_num = 2

    data_start = header_row_num + 1
    current_reg = None

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        values = list(row)
        while len(values) < 12:
            values.append(None)

        offset = str(values[0]).strip() if values[0] else ""
        name = str(values[1]).strip() if values[1] else ""
        bit_range = str(values[5]).strip() if values[5] else ""
        field_name = str(values[6]).strip() if values[6] else ""
        default_val = str(values[7]).strip() if values[7] else ""
        rw_attr = str(values[8]).strip() if values[8] else ""
        desc = str(values[10]).strip() if values[10] else ""

        if offset and name:
            if current_reg:
                registers.append(current_reg)
            current_reg = {
                "offset": offset,
                "name": name,
                "type": str(values[2] or "").strip(),
                "depth": str(values[3] or "").strip(),
                "width": str(values[4] or "").strip(),
                "fields": [],
            }
            if field_name:
                current_reg["fields"].append(
                    {
                        "range": bit_range,
                        "name": field_name,
                        "default": default_val,
                        "access": rw_attr,
                        "description": desc,
                    }
                )
        elif current_reg and field_name:
            current_reg["fields"].append(
                {
                    "range": bit_range,
                    "name": field_name,
                    "default": default_val,
                    "access": rw_attr,
                    "description": desc,
                }
            )

    if current_reg:
        registers.append(current_reg)

    return registers


def parse_xlsx(filepath: str) -> dict[str, Any]:
    """Parse register xlsx into structured JSON.

    Output format matches design spec:
    {
      "source_file": "upa_reg.xlsx",
      "base_addr": "0x0000_1000",
      "total_registers": 150,
      "registers": [...],
      "name_index": {"upa_ctrl": 0, ...}
    }
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Register file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {
        "source_file": filepath.name,
        "base_addr": "0x0000_0000",
        "registers": [],
        "name_index": {},
    }

    # Extract base address from address-map sheet
    if "地址映射" in wb.sheetnames:
        ws = wb["地址映射"]
        for row in ws.iter_rows(values_only=True):
            if row[0] and "base" in str(row[0]).lower():
                result["base_addr"] = (
                    str(row[1]).strip() if row[1] else result["base_addr"]
                )
                break

    skip_sheets = {"封面", "地址映射"}

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        registers = parse_register_sheet(ws)
        for reg in registers:
            idx = len(result["registers"])
            result["name_index"][reg["name"]] = idx
            result["registers"].append(reg)

    result["total_registers"] = len(result["registers"])
    return result


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python reg_to_json.py <input.xlsx> [output.json]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    result = parse_xlsx(input_path)

    print(f"Source: {result['source_file']}")
    print(f"Base address: {result['base_addr']}")
    print(f"Total registers: {result['total_registers']}")

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"Output written to: {output_path}")
