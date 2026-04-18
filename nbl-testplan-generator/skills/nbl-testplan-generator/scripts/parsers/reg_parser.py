#!/usr/bin/env python3
"""reg_parser.py - Parse register configuration .xlsx into structured JSON."""
import json
import sys
from pathlib import Path

import openpyxl

WRITABLE_ACCESS = {"rw", "rww"}


def parse_register_sheet(ws, base_addr="0x0000_0000"):
    """Parse a single register sheet, returning a list of register dicts."""
    registers = []
    name_index = {}
    header_row_num = None

    # Find header row by looking for known column headers
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row[0] and "offset_addr" in str(row[0]).lower():
            header_row_num = i
            break
        if len(row) > 1 and row[1] and "name" in str(row[1]).lower():
            header_row_num = i
            break

    data_start = (header_row_num or 2) + 1

    current_reg = None
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        values = list(row)
        # Pad to at least 12 columns
        while len(values) < 12:
            values.append(None)

        offset = str(values[0]).strip() if values[0] else ""
        name = str(values[1]).strip() if values[1] else ""
        bit_range = str(values[5]).strip() if values[5] else ""
        field_name = str(values[6]).strip() if values[6] else ""
        default_val = str(values[7]).strip() if values[7] else ""
        rw_attr = str(values[8]).strip() if values[8] else ""
        desc = str(values[10]).strip() if values[10] else ""
        config_type = str(values[11]).strip() if values[11] else ""

        if offset and name:
            # New register row
            if current_reg:
                registers.append(current_reg)
            current_reg = {
                "offset": offset,
                "name": name,
                "type": str(values[2] or "").strip(),
                "depth": str(values[3] or "").strip(),
                "width": str(values[4] or "").strip(),
                "fields": [],
                "config_type": config_type,
            }
            if field_name:
                current_reg["fields"].append({
                    "range": bit_range,
                    "name": field_name,
                    "default": default_val,
                    "access": rw_attr,
                    "description": desc,
                    "config_type": config_type,
                })
            name_index[name] = len(registers)
        elif current_reg and field_name:
            # Continuation row (field of current register)
            current_reg["fields"].append({
                "range": bit_range,
                "name": field_name,
                "default": default_val,
                "access": rw_attr,
                "description": desc,
                "config_type": config_type,
            })

    if current_reg:
        registers.append(current_reg)

    return registers, name_index


def parse_xlsx(filepath):
    """Parse a register configuration .xlsx file into a structured dict."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    filepath = Path(filepath)

    result = {"source_file": filepath.name, "sheets": {}}

    # Extract base address from the address-map sheet if present
    base_addr = "0x0000_0000"
    if "\u5730\u5740\u6620\u5c04" in wb.sheetnames:
        ws = wb["\u5730\u5740\u6620\u5c04"]
        for row in ws.iter_rows(values_only=True):
            if row[0] and "base" in str(row[0]).lower():
                base_addr = str(row[1]).strip() if row[1] else base_addr
                break
    result["base_addr"] = base_addr

    # Sheets to skip
    skip_sheets = {"\u5c01\u9762", "\u5730\u5740\u6620\u5c04"}

    all_registers = []
    all_name_index = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        registers, name_index = parse_register_sheet(ws, base_addr)
        result["sheets"][sheet_name] = {
            "register_count": len(registers),
            "registers": registers,
        }
        for name, idx in name_index.items():
            all_name_index[name] = (sheet_name, idx)
        all_registers.extend(registers)

    result["total_registers"] = len(all_registers)
    result["name_index"] = all_name_index
    return result


def filter_writable_fields(parsed):
    """Return a flat list of fields whose access type is rw or rww."""
    writable = []
    for sheet_name, sheet_data in parsed["sheets"].items():
        for reg in sheet_data["registers"]:
            for field in reg["fields"]:
                access = field.get("access", "").lower().strip()
                if access in WRITABLE_ACCESS:
                    writable.append({
                        "register_name": reg["name"],
                        "field_name": field["name"],
                        "range": field.get("range", ""),
                        "default": field.get("default", ""),
                        "access": access,
                        "description": field.get("description", ""),
                        "config_type": field.get("config_type", ""),
                    })
    return writable


def search_registers(parsed, keyword):
    """Search registers by name or field description matching *keyword*."""
    matches = []
    for sheet_name, sheet_data in parsed["sheets"].items():
        for reg in sheet_data["registers"]:
            if keyword.lower() in reg["name"].lower():
                matches.append(reg)
                continue
            for field in reg["fields"]:
                if keyword.lower() in field.get("description", "").lower():
                    matches.append(reg)
                    break
    return matches


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python reg_parser.py <input.xlsx> [output.json]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    result = parse_xlsx(input_path)

    print(f"Source: {result['source_file']}")
    print(f"Base address: {result['base_addr']}")
    print(f"Total registers: {result['total_registers']}")
    for sheet_name, sheet_data in result["sheets"].items():
        print(f"  Sheet '{sheet_name}': {sheet_data['register_count']} registers")

    writable = filter_writable_fields(result)
    print(f"Writable (rw/rww) fields: {len(writable)}")

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"Output written to: {output_path}")
