#!/usr/bin/env python3
"""sample_parser.py - Parse sample test point .xlsx into structured JSON.

Responsibilities:
  - Extract column structure from header rows
  - Collect sample test point entries (up to max_samples)
  - Build feature list with entry counts
  - Provide extract_existing_registers() for incremental append / skip logic
"""
import json
import re
import sys
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_sample_xlsx(filepath, max_samples=20):
    """Parse a sample test-point .xlsx and return a structured result dict.

    Parameters
    ----------
    filepath : str | Path
        Path to the .xlsx file.
    max_samples : int
        Maximum number of sample entries to collect.

    Returns
    -------
    dict with keys:
        source_file, sheet_name, total_rows, total_cols,
        columns, data_start_row, sample_entries, features
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    filepath = Path(filepath)
    ws = wb.active

    result = {
        "source_file": filepath.name,
        "sheet_name": ws.title,
        "total_rows": ws.max_row,
        "total_cols": ws.max_column,
        "columns": [],
        "data_start_row": None,
        "sample_entries": [],
        "features": [],
    }

    # --- Header rows -------------------------------------------------------
    row1 = [
        str(c).strip() if c else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    row2 = [
        str(c).strip() if c else ""
        for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    ]

    # Only keep columns that have meaningful headers (skip trailing None cols)
    meaningful_cols = 0
    for val in row2:
        if val:
            meaningful_cols += 1
        elif meaningful_cols > 0:
            break
    # Ensure we capture at least 19 columns for 19-col format files
    meaningful_cols = max(meaningful_cols, 19)

    for i in range(meaningful_cols):
        r1 = row1[i] if i < len(row1) else ""
        r2 = row2[i] if i < len(row2) else ""
        col_def = {
            "index": i,
            "top_header": r1 if r1 and r1 != "skip" else "",
            "detail_header": r2 if r2 and r2 != "skip" else "",
        }
        col_def["name"] = (
            col_def["detail_header"] or col_def["top_header"] or f"col_{i}"
        )
        result["columns"].append(col_def)

    # --- Detect data start row ---------------------------------------------
    data_start = None
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        values = list(row)
        feature_val = (
            str(values[3]).strip() if len(values) > 3 and values[3] else ""
        )
        subfeature_val = (
            str(values[4]).strip() if len(values) > 4 and values[4] else ""
        )
        if feature_val or subfeature_val:
            data_start = i
            break
    if data_start is None:
        data_start = 4
    result["data_start_row"] = data_start

    # --- Collect sample entries --------------------------------------------
    current_feature = ""
    samples_collected = 0
    feature_entries: dict[str, list] = {}

    for i, row in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True), start=data_start
    ):
        if samples_collected >= max_samples:
            break

        values = list(row)
        # Pad to at least 19 columns
        while len(values) < 19:
            values.append(None)

        def _val(idx):
            return str(values[idx]).strip() if idx < len(values) and values[idx] else ""

        feature = _val(3)
        subfeature = _val(4)
        subfeature2 = _val(5)

        # Track current feature across rows
        if feature:
            current_feature = feature
            if current_feature not in feature_entries:
                feature_entries[current_feature] = []

        # Only collect rows that have subfeature-level content
        if subfeature or subfeature2:
            entry = {
                "row": i,
                "feature": current_feature,
                "subfeature": subfeature,
                "subfeature2": subfeature2,
                "details": _val(9),
                "stimulus": _val(10),
                "checking": _val(11),
                "coverage": _val(12),
                "priority": _val(13),
                "activity": _val(14),
                "path": _val(18),
            }
            result["sample_entries"].append(entry)
            if current_feature in feature_entries:
                feature_entries[current_feature].append(entry)
            samples_collected += 1

    # --- Feature summary ---------------------------------------------------
    result["features"] = [
        {"name": name, "entry_count": len(entries)}
        for name, entries in feature_entries.items()
        if entries  # only include features that have collected entries
    ]
    result["max_data_row"] = ws.max_row

    wb.close()
    return result


def extract_existing_registers(sample_result):
    """Build a mapping of register names found in sample entries.

    Scans the *stimulus* field of each entry for patterns like ``REGNAME.``
    and records the feature/subfeature where each register was seen.

    Parameters
    ----------
    sample_result : dict
        The dict returned by :func:`parse_sample_xlsx`.

    Returns
    -------
    dict[str, str]
        ``{register_name_lower: "feature/subfeature"}``
    """
    reg_map: dict[str, str] = {}
    reg_pattern = re.compile(r"(\w+)\.", re.IGNORECASE)

    for entry in sample_result.get("sample_entries", []):
        feature_name = entry.get("feature", "")
        subfeature = entry.get("subfeature", "")
        stimulus = entry.get("stimulus", "")
        for match in reg_pattern.finditer(stimulus):
            reg_name = match.group(1)
            key = reg_name.lower()
            if key not in reg_map:
                reg_map[key] = f"{feature_name}/{subfeature}"

    return reg_map


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python sample_parser.py <input.xlsx> [output.json]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    result = parse_sample_xlsx(input_path)

    print(f"Source: {result['source_file']}")
    print(f"Sheet:  {result['sheet_name']}")
    print(f"Size:   {result['total_rows']} rows x {result['total_cols']} cols")
    print(f"Data starts at row: {result['data_start_row']}")
    print(f"Features: {[f['name'] for f in result['features']]}")
    print(f"Sample entries collected: {len(result['sample_entries'])}")

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"Output written to: {output_path}")
