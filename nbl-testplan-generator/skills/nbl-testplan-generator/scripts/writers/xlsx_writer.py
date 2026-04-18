"""xlsx_writer.py -- Write hierarchical testpoint data into an xlsx test plan.

Supports Chapter 1 (functional features) and Chapter 2 (register configuration
features) with outline grouping, hidden columns, freeze panes, auto-filter,
and wrap-text formatting.
"""

from __future__ import annotations

import copy
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline

DEFAULT_FONT_NAME = "微软雅黑"
DEFAULT_FONT_SIZE = 10
DEFAULT_FONT_COLOR = "000000"  # black
RED_FONT_COLOR = "FF0000"

# Column indices (1-based) that should be hidden in the output.
HIDDEN_COLUMNS: list[int] = [3, 14, 15, 16, 17, 18, 19, 20]

# Columns that need wrap_text on their data cells.
WRAP_TEXT_COLUMNS: set[int] = set(range(4, 24))  # D through W — all data columns

# Thin border for all non-empty data cells
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Gold light 80% fill for cells containing ⚠ or [推断]
GOLD_LIGHT_FILL = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid",
)


def _apply_meta_rows(ws: openpyxl.worksheet.worksheet.Worksheet, spec_name: str) -> None:
    """Write rows 3-5 metadata: plan / weight / skip rows."""
    default_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, color=DEFAULT_FONT_COLOR)
    c = ws.cell(row=3, column=1, value="plan")
    c.font = default_font
    c = ws.cell(row=4, column=1, value="weight")
    c.font = default_font
    c = ws.cell(row=4, column=21, value=5)
    c.font = default_font
    c = ws.cell(row=4, column=22, value=95)
    c.font = default_font
    c = ws.cell(row=5, column=1, value="skip")
    c.font = default_font
    c = ws.cell(row=5, column=2, value=spec_name)
    c.font = default_font


def _apply_sheet_formatting(ws: openpyxl.worksheet.worksheet.Worksheet, last_row: int) -> None:
    """Apply outline properties, hidden columns, freeze panes, auto-filter."""
    # Outline: summaryBelow=False so grouping buttons appear above groups.
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False)

    # Hide specific columns.
    for col in HIDDEN_COLUMNS:
        ws.column_dimensions[get_column_letter(col)].hidden = True

    # Freeze panes at A3.
    ws.freeze_panes = "A3"

    # Auto-filter on A2:W{last_row}.
    ws.auto_filter.ref = f"A2:W{last_row}"


def _apply_rectangle_borders(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
) -> None:
    """Apply thin borders to the full A-W rectangle for data rows."""
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, 24):  # A through W
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER


BREAK_AFTER_CHARS = [";", "；", "。", "：", ":", "（", "）", "(", ")"]


def _insert_line_breaks(text: str) -> str:
    """Insert line-breaks after punctuation characters."""
    if not text:
        return text
    result = []
    i = 0
    while i < len(text):
        ch = text[i]
        result.append(ch)
        if ch in BREAK_AFTER_CHARS:
            if i + 1 < len(text) and text[i + 1] != "\n":
                result.append("\n")
        i += 1
    return "".join(result)


def _write_cell(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int,
    value: Any,
) -> None:
    """Write a data cell with standard formatting.

    - Non-empty cells: left-align, vertical-center, wrap_text
    - Insert line-breaks after punctuation (except W column)
    - Gold background on cells containing '⚠' or '[推断]'
    - Borders applied later via _apply_rectangle_borders
    """
    if value is None or (isinstance(value, str) and value.strip() == ""):
        cell = ws.cell(row=row, column=col)
        cell.value = None
        # Clear any template-inherited borders on empty cells
        cell.border = Border()
        return

    cell = ws.cell(row=row, column=col, value=value)
    text = str(value)
    # Insert line-breaks after punctuation (except W column which uses commas)
    if col != 23:
        text = _insert_line_breaks(text)
        cell.value = text
    # Sanitize W column: replace newlines with commas
    if col == 23 and "\n" in text:
        text = ",".join(part.strip() for part in text.split("\n") if part.strip())
        cell.value = text
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    cell.font = Font(
        name=DEFAULT_FONT_NAME,
        size=DEFAULT_FONT_SIZE,
        color=DEFAULT_FONT_COLOR,
    )

    # Red font for I-column cells containing 【配置】 or 【激励】
    if col == 9 and ("【配置】" in text or "【激励】" in text):
        cell.font = Font(
            name=DEFAULT_FONT_NAME,
            size=DEFAULT_FONT_SIZE,
            color=RED_FONT_COLOR,
        )

    # Conditional gold fill for inferred/uncertain content
    if "⚠" in text or "[推断]" in text:
        cell.fill = GOLD_LIGHT_FILL
        cell.font = Font(
            name=DEFAULT_FONT_NAME,
            size=DEFAULT_FONT_SIZE,
            color=RED_FONT_COLOR,
        )


def _write_ch1(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    data: dict[str, Any],
    start_row: int,
) -> tuple[int, dict[str, int]]:
    """Write Chapter 1 (functional features) rows.

    Returns (next_free_row, stats_partial).
    """
    row = start_row
    total_features = 0
    total_l1 = 0
    total_l2 = 0

    chapter_label = data.get("chapter", "chapter 1 功能特性")

    for i, feat in enumerate(data.get("features", [])):
        total_features += 1
        feature_name = feat.get("feature", "")
        feature_id = feat.get("feature_id", "")
        feature_chapter = feat.get("chapter", chapter_label)

        # Feature header row: D = feature name (outline_level 0)
        # Only first feature row gets the chapter label in B column
        if i == 0:
            _write_cell(ws, row, 2, chapter_label)
        _write_cell(ws, row, 4, feature_name)
        ws.row_dimensions[row].outline_level = 0
        row += 1

        for sf1 in feat.get("subfeatures_l1", []):
            total_l1 += 1
            sf1_name = sf1.get("subfeature_l1", "")

            # L1 row: E = subfeature_l1 (outline_level 1)
            _write_cell(ws, row, 5, sf1_name)
            ws.row_dimensions[row].outline_level = 1
            row += 1

            for sf2 in sf1.get("subfeatures_l2", []):
                total_l2 += 1
                # L2 row: F/G outline_level 2, then D-H, I, J, K, L, M, W
                overview = sf2.get("subfeature_l2_overview", "")
                detail = sf2.get("subfeature_l2_detail", "")
                remarks = sf2.get("remarks", "")
                stimulus = sf2.get("stimulus", "")
                checking = sf2.get("checking", "")
                coverage = sf2.get("coverage", "")
                priority = sf2.get("priority", "")
                activity = sf2.get("activity", "")
                path = sf2.get("path", "")

                _write_cell(ws, row, 6, overview)
                _write_cell(ws, row, 7, detail)
                _write_cell(ws, row, 8, remarks)
                _write_cell(ws, row, 9, stimulus)
                _write_cell(ws, row, 10, checking)
                _write_cell(ws, row, 11, coverage)
                _write_cell(ws, row, 12, priority)
                _write_cell(ws, row, 13, activity)
                _write_cell(ws, row, 23, path)

                ws.row_dimensions[row].outline_level = 2
                row += 1

    return row, {
        "total_features": total_features,
        "total_l1": total_l1,
        "total_l2": total_l2,
        "skipped_regs": 0,
    }


def _write_ch2(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    data: dict[str, Any],
    start_row: int,
) -> tuple[int, dict[str, int]]:
    """Write Chapter 2 (register configuration features) rows.

    Returns (next_free_row, stats_partial).
    """
    row = start_row
    total_features = 0
    total_l1 = 0
    total_l2 = 0
    skipped_regs = 0

    chapter_label = data.get("chapter", "chapter 2 配置特性")

    for i, reg in enumerate(data.get("registers", [])):
        reg_name = reg.get("register_name", "")
        total_features += 1

        # Register header row: D = register name (outline_level 0)
        if i == 0:
            _write_cell(ws, row, 2, chapter_label)
        _write_cell(ws, row, 4, reg_name)
        ws.row_dimensions[row].outline_level = 0
        row += 1

        for sf1 in reg.get("subfeatures_l1", []):
            total_l1 += 1
            sf1_name = sf1.get("subfeature_l1", "")

            # L1 row: E = subfeature_l1 (outline_level 1)
            _write_cell(ws, row, 5, sf1_name)
            ws.row_dimensions[row].outline_level = 1
            row += 1

            for sf2 in sf1.get("subfeatures_l2", []):
                total_l2 += 1
                is_skip = sf2.get("skip", False)

                overview = sf2.get("subfeature_l2_overview", "")
                detail = sf2.get("subfeature_l2_detail", "")
                remarks = sf2.get("remarks", "")
                stimulus = sf2.get("stimulus", "")
                checking = sf2.get("checking", "")
                coverage = sf2.get("coverage", "")
                priority = sf2.get("priority", "")
                activity = sf2.get("activity", "")
                path = sf2.get("path", "")

                # If skip=True: A = "skip"
                if is_skip:
                    _write_cell(ws, row, 1, "skip")
                    skipped_regs += 1

                _write_cell(ws, row, 6, overview)
                _write_cell(ws, row, 7, detail)
                _write_cell(ws, row, 8, remarks)
                _write_cell(ws, row, 9, stimulus)
                _write_cell(ws, row, 10, checking)
                _write_cell(ws, row, 11, coverage)
                _write_cell(ws, row, 12, priority)
                _write_cell(ws, row, 13, activity)
                _write_cell(ws, row, 23, path)

                ws.row_dimensions[row].outline_level = 2
                row += 1

    return row, {
        "total_features": total_features,
        "total_l1": total_l1,
        "total_l2": total_l2,
        "skipped_regs": skipped_regs,
    }


def write_testpoint_xlsx(
    chapters: list[dict[str, Any]],
    output_path: str,
    template_path: str,
) -> dict[str, int]:
    """Write hierarchical testpoint data into an xlsx test plan.

    Parameters
    ----------
    chapters : list of dict
        Each dict has ``"type"`` (``"ch1"`` or ``"ch2"``) and ``"data"``.
    output_path : str
        Destination xlsx file path.
    template_path : str
        Path to the template xlsx to clone formatting/headers from.

    Returns
    -------
    dict[str, int]
        Stats: ``total_features``, ``total_l1``, ``total_l2``, ``skipped_regs``.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Determine spec_name from the first ch1 chapter if available.
    spec_name = ""
    for ch in chapters:
        if ch["type"] == "ch1":
            spec_name = ch["data"].get("spec_name", "")
            break

    # Clear template example data rows (row 6 onwards).
    max_row = ws.max_row
    for row_idx in range(6, max_row + 1):
        for col_idx in range(1, 24):
            ws.cell(row=row_idx, column=col_idx).value = None
            ws.cell(row=row_idx, column=col_idx).border = Border()
        ws.row_dimensions[row_idx].outline_level = 0

    # Write meta rows (3-5).
    _apply_meta_rows(ws, spec_name)

    # Write chapter data starting at row 6.
    row = 6
    stats: dict[str, int] = {
        "total_features": 0,
        "total_l1": 0,
        "total_l2": 0,
        "skipped_regs": 0,
    }

    for ch in chapters:
        if ch["type"] == "ch1":
            row, ch_stats = _write_ch1(ws, ch["data"], row)
        elif ch["type"] == "ch2":
            row, ch_stats = _write_ch2(ws, ch["data"], row)
        else:
            continue

        for key in stats:
            stats[key] += ch_stats.get(key, 0)

    last_row = max(row - 1, 6)

    # Apply rectangle borders to all data rows
    data_start = 6
    data_end = max(row - 1, 6)
    _apply_rectangle_borders(ws, data_start, data_end)

    # Apply sheet-level formatting.
    _apply_sheet_formatting(ws, last_row)

    wb.save(output_path)
    wb.close()

    return stats
