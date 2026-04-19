#!/usr/bin/env python3
"""reg_to_json.py - Convert register xlsx to structured JSON."""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import openpyxl


def parse_register_xlsx(filepath: str | Path) -> dict[str, Any]:
    """Parse register xlsx into structured JSON."""
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"Register file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    result: dict[str, Any] = {
        "source_file": filepath.name,
        "total_registers": 0,
        "registers": [],
        "name_index": {},
    }

    skip_sheets = {"封面", "地址映射", "address_map", "summary"}

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in {s.lower() for s in skip_sheets}:
            continue

        ws = wb[sheet_name]
        registers = _parse_register_sheet(ws)

        for reg in registers:
            idx = len(result["registers"])
            result["name_index"][reg["name"]] = idx
            result["registers"].append(reg)

    result["total_registers"] = len(result["registers"])
    wb.close()
    return result


def _parse_register_sheet(ws) -> list[dict[str, Any]]:
    """Parse a single register sheet."""
    registers: list[dict[str, Any]] = []

    # Find header row
    header_row_num = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        row_values = [str(v).lower() if v else "" for v in row]
        if any("offset" in v for v in row_values) or any("name" in v for v in row_values[:3]):
            header_row_num = i
            break

    if header_row_num is None:
        return registers

    data_start = header_row_num + 1
    current_reg: dict[str, Any] | None = None

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        values = list(row)
        while len(values) < 11:
            values.append(None)

        offset = str(values[0]).strip() if values[0] else ""
        name = str(values[1]).strip() if values[1] else ""
        bit_range = str(values[5]).strip() if len(values) > 5 and values[5] else ""
        field_name = str(values[6]).strip() if len(values) > 6 and values[6] else ""
        default_val = str(values[7]).strip() if len(values) > 7 and values[7] else ""
        rw_attr = str(values[8]).strip() if len(values) > 8 and values[8] else ""
        desc = str(values[10]).strip() if len(values) > 10 and values[10] else ""

        if offset and name:
            if current_reg:
                registers.append(current_reg)

            current_reg = {
                "offset": offset,
                "name": name,
                "type": str(values[2] or "").strip() if len(values) > 2 else "",
                "fields": [],
            }

            if field_name:
                current_reg["fields"].append({
                    "range": bit_range,
                    "name": field_name,
                    "default": default_val,
                    "access": rw_attr,
                    "description": desc,
                })

        elif current_reg and field_name:
            current_reg["fields"].append({
                "range": bit_range,
                "name": field_name,
                "default": default_val,
                "access": rw_attr,
                "description": desc,
            })

    if current_reg:
        registers.append(current_reg)

    return registers


def get_register_by_name(reg_data: dict[str, Any], name: str) -> dict[str, Any] | None:
    """Get register by name from parsed register data."""
    idx = reg_data.get("name_index", {}).get(name)
    if idx is not None:
        return reg_data["registers"][idx]
    return None


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python reg_to_json.py <input.xlsx> [output.json]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    result = parse_register_xlsx(input_path)

    print(f"Source: {result['source_file']}")
    print(f"Total registers: {result['total_registers']}")
    print(f"Registers: {[r['name'] for r in result['registers']]}")

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"Output written to: {output_path}")
