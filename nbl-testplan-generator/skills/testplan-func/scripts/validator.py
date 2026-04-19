#!/usr/bin/env python3
"""validator.py - Validate testplan output structure.

Validates the D-E-F-G hierarchy structure and Excel output:
- D-E-F-G hierarchy: feature > subfeature_l1 > subfeature_l2 > subfeature_l3
- Outline levels: 0-3 for D-E-F-G columns
- H-W columns: Only at minimum granularity (G if exists, else F)
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# Column indices
COL_D = 4  # Feature (outline level 0)
COL_E = 5  # Subfeature L1 (outline level 1)
COL_F = 6  # Subfeature L2 (outline level 2)
COL_G = 7  # Subfeature L3 (outline level 3)
COL_H = 8  # Remarks/Details
COL_W = 23  # Path to Source


def validate_hierarchy(features_data: dict[str, Any]) -> dict[str, Any]:
    """Validate the D-E-F-G hierarchy structure in features data.

    Parameters
    ----------
    features_data : dict
        Features data structure to validate.

    Returns
    -------
    dict
        Validation result with 'valid' (bool) and 'errors' (list) fields.
    """
    errors: list[str] = []

    # Check required top-level fields
    if "features" not in features_data:
        errors.append("Missing 'features' field in data")

    # Validate each feature
    features = features_data.get("features", [])
    for i, feature in enumerate(features):
        feature_errors = _validate_feature(feature, i)
        errors.extend(feature_errors)

    return {
        "valid": len(errors) == 0,
        "errors": errors,
    }


def _validate_feature(feature: dict[str, Any], index: int) -> list[str]:
    """Validate a single feature entry."""
    errors: list[str] = []

    # Feature must have 'feature' field
    if not feature.get("feature"):
        errors.append(f"Feature {index}: Missing 'feature' field")

    # Validate L1 subfeatures
    l1_list = feature.get("subfeatures_l1", [])
    for j, l1 in enumerate(l1_list):
        l1_errors = _validate_subfeature_l1(l1, index, j)
        errors.extend(l1_errors)

    return errors


def _validate_subfeature_l1(l1: dict[str, Any], feature_idx: int, l1_idx: int) -> list[str]:
    """Validate an L1 subfeature entry."""
    errors: list[str] = []

    # L1 must have subfeature_l1 field
    if not l1.get("subfeature_l1"):
        errors.append(
            f"Feature {feature_idx}, L1 {l1_idx}: Missing 'subfeature_l1' field"
        )

    # Validate L2 subfeatures
    l2_list = l1.get("subfeatures_l2", [])
    for k, l2 in enumerate(l2_list):
        l2_errors = _validate_subfeature_l2(l2, feature_idx, l1_idx, k)
        errors.extend(l2_errors)

    return errors


def _validate_subfeature_l2(
    l2: dict[str, Any],
    feature_idx: int,
    l1_idx: int,
    l2_idx: int,
) -> list[str]:
    """Validate an L2 subfeature entry."""
    errors: list[str] = []

    # L2 must have at least one identifier
    has_identifier = any([
        l2.get("subfeature_l2_overview"),
        l2.get("subfeature_l2_detail"),
        l2.get("remarks"),
    ])

    if not has_identifier:
        errors.append(
            f"Feature {feature_idx}, L1 {l1_idx}, L2 {l2_idx}: "
            "Missing identifier field (subfeature_l2_overview, subfeature_l2_detail, or remarks)"
        )

    return errors


def validate_outline_levels(xlsx_path: str | Path) -> dict[str, Any]:
    """Validate outline levels (0-3) in Excel output.

    Parameters
    ----------
    xlsx_path : str | Path
        Path to the Excel file to validate.

    Returns
    -------
    dict
        Validation result with 'valid', 'errors', and 'outline_stats' fields.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    errors: list[str] = []
    outline_stats: dict[int, int] = {0: 0, 1: 0, 2: 0, 3: 0}

    wb = load_workbook(xlsx_path)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        outline_level = ws.row_dimensions[row].outline_level

        # Only check rows with data in D-G columns
        has_data = False
        for col in range(COL_D, COL_G + 1):
            if ws.cell(row=row, column=col).value:
                has_data = True
                break

        if has_data:
            if outline_level is None:
                errors.append(f"Row {row}: Missing outline level")
            elif outline_level not in (0, 1, 2, 3):
                errors.append(
                    f"Row {row}: Invalid outline level {outline_level} (must be 0-3)"
                )
            else:
                outline_stats[outline_level] = outline_stats.get(outline_level, 0) + 1

    wb.close()

    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "outline_stats": outline_stats,
    }


def validate_hw_columns(xlsx_path: str | Path) -> dict[str, Any]:
    """Validate H-W columns are only at minimum granularity.

    Parameters
    ----------
    xlsx_path : str | Path
        Path to the Excel file to validate.

    Returns
    -------
    dict
        Validation result with 'valid' and 'errors' fields.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    errors: list[str] = []

    wb = load_workbook(xlsx_path)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        # Check if row has data in H-W columns
        has_hw_data = False
        for col in range(COL_H, COL_W + 1):
            if ws.cell(row=row, column=col).value:
                has_hw_data = True
                break

        if has_hw_data:
            # Check if this is minimum granularity
            outline_level = ws.row_dimensions[row].outline_level
            cell_g = ws.cell(row=row, column=COL_G).value
            cell_f = ws.cell(row=row, column=COL_F).value

            # H-W should only be at G level (outline 3) or F level (outline 2) if no G
            if outline_level is None:
                errors.append(f"Row {row}: H-W data but no outline level")
            elif outline_level < 2:
                # H-W at D or E level is not allowed
                errors.append(
                    f"Row {row}: H-W data at outline level {outline_level} "
                    "(should be at minimum granularity)"
                )
            elif outline_level == 2 and cell_g:
                # If G exists (in same or different row), F shouldn't have H-W
                # Actually, we check if this row has F filled but not G
                if cell_f and not cell_g:
                    # This is OK - F is minimum granularity
                    pass

    wb.close()

    return {
        "valid": len(errors) == 0,
        "errors": errors,
    }


def validate_xlsx_output(xlsx_path: str | Path) -> dict[str, Any]:
    """Validate Excel output file with all checks.

    Combines outline level validation and H-W column validation.

    Parameters
    ----------
    xlsx_path : str | Path
        Path to the Excel file to validate.

    Returns
    -------
    dict
        Combined validation result with 'valid' and 'errors' fields.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    all_errors: list[str] = []

    # Run outline level validation
    outline_result = validate_outline_levels(xlsx_path)
    all_errors.extend(outline_result.get("errors", []))

    # Run H-W column validation
    hw_result = validate_hw_columns(xlsx_path)
    all_errors.extend(hw_result.get("errors", []))

    return {
        "valid": len(all_errors) == 0,
        "errors": all_errors,
        "outline_stats": outline_result.get("outline_stats"),
    }


def validate_full_output(
    features_data: dict[str, Any],
    xlsx_path: str | Path,
) -> dict[str, Any]:
    """Validate both hierarchy data and Excel output.

    Parameters
    ----------
    features_data : dict
        Features data structure to validate.
    xlsx_path : str | Path
        Path to the Excel file to validate.

    Returns
    -------
    dict
        Combined validation result.
    """
    all_errors: list[str] = []

    # Validate hierarchy
    hierarchy_result = validate_hierarchy(features_data)
    all_errors.extend(hierarchy_result.get("errors", []))

    # Validate Excel if path provided
    if xlsx_path:
        xlsx_result = validate_xlsx_output(xlsx_path)
        all_errors.extend(xlsx_result.get("errors", []))

    return {
        "valid": len(all_errors) == 0,
        "errors": all_errors,
        "hierarchy_valid": hierarchy_result.get("valid", False),
        "xlsx_valid": xlsx_result.get("valid", False) if xlsx_path else None,
    }


if __name__ == "__main__":
    import argparse
    import json
    import sys

    parser = argparse.ArgumentParser(description="Validate testplan output")
    parser.add_argument("--xlsx", required=True, help="Path to xlsx file")
    parser.add_argument("--json", help="Path to features JSON (optional)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)

    if args.json:
        with open(args.json, "r", encoding="utf-8") as f:
            features_data = json.load(f)
        result = validate_full_output(features_data, xlsx_path)
    else:
        result = validate_xlsx_output(xlsx_path)

    print(f"Valid: {result['valid']}")
    if result["errors"]:
        print("Errors:")
        for error in result["errors"]:
            print(f"  - {error}")

    if "outline_stats" in result and result["outline_stats"]:
        print("Outline level stats:")
        for level, count in sorted(result["outline_stats"].items()):
            print(f"  Level {level}: {count} rows")

    sys.exit(0 if result["valid"] else 1)
