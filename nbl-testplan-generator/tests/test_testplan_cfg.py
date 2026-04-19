"""Tests for cfg_writer module - Configuration Feature Generator."""
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Add scripts to path
sys.path.insert(0, str(Path(__file__).parent.parent / "skills" / "testplan-cfg" / "scripts"))

from cfg_writer import (
    CFG_EXCLUDE_PATTERNS,
    VALID_CFG_ACCESS,
    _is_excluded_register,
    append_cfg_to_func_xlsx,
    build_ch2_json,
    cross_ref_cfg,
    extract_covered_regs_from_xlsx,
    filter_cfg_registers,
    group_similar_registers,
    write_independent_ch2_xlsx,
)
from reg_to_json import parse_register_xlsx


class TestFilterCfgRegisters:
    """Tests for filter_cfg_registers function."""

    def test_filter_cfg_registers_exists(self):
        """filter_cfg_registers function should exist."""
        assert callable(filter_cfg_registers)

    def test_filter_excludes_int_registers(self):
        """Should exclude *_int_* registers."""
        registers = [
            {"name": "upa_ctrl", "fields": [{"name": "en", "access": "RW"}]},
            {"name": "upa_int_status", "fields": [{"name": "irq", "access": "RW"}]},
            {"name": "upa_int_mask", "fields": [{"name": "mask", "access": "RW"}]},
        ]
        result = filter_cfg_registers(registers)
        names = [r["name"] for r in result]
        assert "upa_ctrl" in names
        assert "upa_int_status" not in names
        assert "upa_int_mask" not in names

    def test_filter_excludes_fifo_registers(self):
        """Should exclude *_fifo_* registers."""
        registers = [
            {"name": "upa_ctrl", "fields": [{"name": "en", "access": "RW"}]},
            {"name": "upa_fifo_th", "fields": [{"name": "th", "access": "RW"}]},
            {"name": "upa_fifo_full", "fields": [{"name": "full", "access": "RO"}]},
        ]
        result = filter_cfg_registers(registers)
        names = [r["name"] for r in result]
        assert "upa_ctrl" in names
        assert "upa_fifo_th" not in names
        assert "upa_fifo_full" not in names

    def test_filter_keeps_only_rw_fields(self):
        """Should keep only RW/RWC/WO/RWW fields."""
        registers = [
            {
                "name": "upa_ctrl",
                "fields": [
                    {"name": "en", "access": "RW"},
                    {"name": "status", "access": "RO"},
                    {"name": "irq", "access": "RWC"},
                ],
            },
        ]
        result = filter_cfg_registers(registers)
        assert len(result) == 1
        field_names = [f["name"] for f in result[0]["fields"]]
        assert "en" in field_names
        assert "status" not in field_names
        assert "irq" in field_names


class TestGroupSimilarRegisters:
    """Tests for group_similar_registers function."""

    def test_group_similar_registers_exists(self):
        """group_similar_registers function should exist."""
        assert callable(group_similar_registers)

    def test_group_merges_numbered_registers(self):
        """Should merge registers with trailing numbers."""
        registers = [
            {"name": "upa_stage_0", "offset": "0x00", "fields": [{"name": "cfg", "access": "RW"}]},
            {"name": "upa_stage_1", "offset": "0x04", "fields": [{"name": "cfg", "access": "RW"}]},
            {"name": "upa_stage_2", "offset": "0x08", "fields": [{"name": "cfg", "access": "RW"}]},
        ]
        result = group_similar_registers(registers)
        assert len(result) == 1
        assert result[0]["name"] == "upa_stage_0/1/2"

    def test_group_keeps_single_registers(self):
        """Should keep single registers unchanged."""
        registers = [
            {"name": "upa_ctrl", "offset": "0x00", "fields": [{"name": "en", "access": "RW"}]},
        ]
        result = group_similar_registers(registers)
        assert len(result) == 1
        assert result[0]["name"] == "upa_ctrl"

    def test_group_merges_fields(self):
        """Should merge fields from all grouped registers."""
        registers = [
            {"name": "upa_ch_0", "fields": [{"name": "en", "access": "RW"}]},
            {"name": "upa_ch_1", "fields": [{"name": "mode", "access": "RW"}]},
        ]
        result = group_similar_registers(registers)
        field_names = [f["name"] for f in result[0]["fields"]]
        assert "en" in field_names
        assert "mode" in field_names


class TestCrossRefCfg:
    """Tests for cross_ref_cfg function."""

    def test_cross_ref_cfg_exists(self):
        """cross_ref_cfg function should exist."""
        assert callable(cross_ref_cfg)

    def test_cross_ref_marks_covered_as_skip(self):
        """Should mark covered registers as skip."""
        registers = [
            {"name": "upa_ctrl", "fields": [{"name": "en", "access": "RW", "range": "0", "description": "Enable"}]},
        ]
        covered_regs = {"upa_ctrl"}
        result = cross_ref_cfg(registers, covered_regs)
        assert len(result) == 1
        assert result[0]["skip"] is True
        assert "覆盖" in result[0]["remarks"]

    def test_cross_ref_generates_stimulus_for_uncovered(self):
        """Should generate stimulus for uncovered registers."""
        registers = [
            {
                "name": "upa_ctrl",
                "fields": [
                    {"name": "en", "access": "RW", "range": "0", "description": "Enable"},
                ],
            },
        ]
        covered_regs = set()
        result = cross_ref_cfg(registers, covered_regs)
        assert len(result) == 1
        assert result[0]["skip"] is False
        assert "【配置】" in result[0]["stimulus"]

    def test_cross_ref_handles_grouped_names(self):
        """Should handle grouped register names."""
        registers = [
            {"name": "upa_ch_0/1", "fields": [{"name": "en", "access": "RW", "range": "0", "description": "Enable"}]},
        ]
        covered_regs = {"upa_ch_0"}
        result = cross_ref_cfg(registers, covered_regs)
        assert result[0]["skip"] is True


class TestBuildCh2Json:
    """Tests for build_ch2_json function."""

    def test_build_ch2_json_exists(self):
        """build_ch2_json function should exist."""
        assert callable(build_ch2_json)

    def test_build_ch2_json_returns_dict(self, sample_reg_xlsx):
        """build_ch2_json should return a dictionary."""
        reg_data = parse_register_xlsx(sample_reg_xlsx)
        result = build_ch2_json(reg_data)
        assert isinstance(result, dict)

    def test_build_ch2_json_has_chapter(self, sample_reg_xlsx):
        """build_ch2_json should include chapter field."""
        reg_data = parse_register_xlsx(sample_reg_xlsx)
        result = build_ch2_json(reg_data)
        assert "chapter" in result
        assert "chapter 2" in result["chapter"].lower()

    def test_build_ch2_json_has_registers(self, sample_reg_xlsx):
        """build_ch2_json should include registers list."""
        reg_data = parse_register_xlsx(sample_reg_xlsx)
        result = build_ch2_json(reg_data)
        assert "registers" in result
        assert isinstance(result["registers"], list)


class TestWriteIndependentCh2Xlsx:
    """Tests for write_independent_ch2_xlsx function."""

    def test_write_independent_ch2_xlsx_exists(self):
        """write_independent_ch2_xlsx function should exist."""
        assert callable(write_independent_ch2_xlsx)

    def test_write_independent_creates_file(self, temp_dir, sample_reg_xlsx, sample_template_xlsx):
        """write_independent_ch2_xlsx should create an output xlsx file."""
        reg_data = parse_register_xlsx(sample_reg_xlsx)
        output_path = temp_dir / "ch2_output.xlsx"
        write_independent_ch2_xlsx(reg_data, str(output_path), str(sample_template_xlsx))
        assert output_path.exists()

    def test_write_independent_returns_stats(self, temp_dir, sample_reg_xlsx, sample_template_xlsx):
        """write_independent_ch2_xlsx should return statistics."""
        reg_data = parse_register_xlsx(sample_reg_xlsx)
        output_path = temp_dir / "ch2_output.xlsx"
        stats = write_independent_ch2_xlsx(reg_data, str(output_path), str(sample_template_xlsx))
        assert "total_registers" in stats
        assert "new_registers" in stats


class TestAppendCfgToFuncXlsx:
    """Tests for append_cfg_to_func_xlsx function."""

    def test_append_cfg_to_func_xlsx_exists(self):
        """append_cfg_to_func_xlsx function should exist."""
        assert callable(append_cfg_to_func_xlsx)

    def test_append_cfg_creates_output(self, temp_dir, sample_reg_xlsx, sample_template_xlsx):
        """append_cfg_to_func_xlsx should create output file."""
        reg_data = parse_register_xlsx(sample_reg_xlsx)
        # First create a func xlsx with minimal data
        func_path = temp_dir / "func.xlsx"
        output_path = temp_dir / "combined.xlsx"

        # Create a minimal func xlsx from template
        wb = load_workbook(sample_template_xlsx)
        ws = wb.active
        # Add minimal Ch1 data
        ws.cell(row=6, column=4, value="Test Feature")
        ws.row_dimensions[6].outline_level = 0
        wb.save(func_path)
        wb.close()

        stats = append_cfg_to_func_xlsx(str(func_path), reg_data, str(output_path))
        assert output_path.exists()

    def test_append_cfg_returns_stats(self, temp_dir, sample_reg_xlsx, sample_template_xlsx):
        """append_cfg_to_func_xlsx should return statistics."""
        reg_data = parse_register_xlsx(sample_reg_xlsx)
        func_path = temp_dir / "func.xlsx"
        output_path = temp_dir / "combined.xlsx"

        # Create a minimal func xlsx from template
        wb = load_workbook(sample_template_xlsx)
        ws = wb.active
        ws.cell(row=6, column=4, value="Test Feature")
        ws.row_dimensions[6].outline_level = 0
        wb.save(func_path)
        wb.close()

        stats = append_cfg_to_func_xlsx(str(func_path), reg_data, str(output_path))
        assert "total_registers" in stats
        assert "covered_registers" in stats
        assert "new_registers" in stats


class TestExtractCoveredRegsFromXlsx:
    """Tests for extract_covered_regs_from_xlsx function."""

    def test_extract_covered_regs_exists(self):
        """extract_covered_regs_from_xlsx function should exist."""
        assert callable(extract_covered_regs_from_xlsx)

    def test_extract_finds_register_in_stimulus(self, temp_dir, sample_template_xlsx):
        """Should find register names in stimulus column."""
        # Create a test xlsx with register reference
        wb = load_workbook(sample_template_xlsx)
        ws = wb.active

        # Add header row
        for row in range(1, 6):
            ws.cell(row=row, column=9, value="stimulus" if row == 2 else None)

        # Add data with register reference
        ws.cell(row=7, column=9, value="【配置】配置 upa_ctrl.en (0)")
        ws.cell(row=8, column=9, value="【配置】配置 upa_mode.mode (1)")

        test_path = temp_dir / "test_func.xlsx"
        wb.save(test_path)
        wb.close()

        covered = extract_covered_regs_from_xlsx(str(test_path))
        assert "upa_ctrl" in covered
        assert "upa_mode" in covered


class TestIsExcludedRegister:
    """Tests for _is_excluded_register function."""

    def test_is_excluded_register_exists(self):
        """_is_excluded_register function should exist."""
        assert callable(_is_excluded_register)

    def test_excludes_int_status(self):
        """Should exclude *_int_status registers."""
        assert _is_excluded_register("upa_int_status") is True

    def test_excludes_fifo_registers(self):
        """Should exclude *_fifo_* registers."""
        assert _is_excluded_register("upa_fifo_full") is True
        assert _is_excluded_register("upa_fifo_th") is True

    def test_keeps_normal_registers(self):
        """Should keep normal configuration registers."""
        assert _is_excluded_register("upa_ctrl") is False
        assert _is_excluded_register("upa_mode") is False
