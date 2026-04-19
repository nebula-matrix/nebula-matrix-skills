"""Tests for xlsx_writer module."""
import importlib.util
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Explicitly load xlsx_writer from testplan-func/scripts
_func_xlsx_writer = Path(__file__).parent.parent / "skills" / "testplan-func" / "scripts" / "xlsx_writer.py"
_spec = importlib.util.spec_from_file_location("xlsx_writer_func", _func_xlsx_writer)
xlsx_writer_func = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(xlsx_writer_func)

write_ch1_xlsx = xlsx_writer_func.write_ch1_xlsx


class TestWriteCh1Xlsx:
    """Tests for write_ch1_xlsx function."""

    def test_write_ch1_xlsx_creates_file(self, temp_dir, sample_template_xlsx):
        """write_ch1_xlsx should create an output xlsx file."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)
        assert output_path.exists()

    def test_write_ch1_xlsx_returns_path(self, temp_dir, sample_template_xlsx):
        """write_ch1_xlsx should return the output path."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [],
        }
        output_path = temp_dir / "output.xlsx"
        result = write_ch1_xlsx(features_data, sample_template_xlsx, output_path)
        assert result == output_path

    def test_write_ch1_xlsx_with_single_feature(self, temp_dir, sample_template_xlsx):
        """write_ch1_xlsx should write a single feature correctly."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "chapter": "chapter 1",
                    "subfeatures_l1": [],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        # Find the feature row (should have value in D column)
        found_feature = False
        for row in range(1, ws.max_row + 1):
            cell_d = ws.cell(row=row, column=4)  # Column D
            if cell_d.value and "Data Path" in str(cell_d.value):
                found_feature = True
                break
        assert found_feature, "Feature not found in output"

    def test_outline_levels_d_is_zero(self, temp_dir, sample_template_xlsx):
        """Feature column D should have outline level 0."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "chapter": "chapter 1",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "Transfer",
                            "subfeatures_l2": [],
                        }
                    ],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Find row with Feature in D (outline level should be 0)
        for row in range(1, ws.max_row + 1):
            cell_d = ws.cell(row=row, column=4)
            if cell_d.value and "Data Path" in str(cell_d.value):
                outline_level = ws.row_dimensions[row].outline_level
                assert outline_level == 0, f"D column row {row} should have outline level 0, got {outline_level}"
                break

    def test_outline_levels_e_is_one(self, temp_dir, sample_template_xlsx):
        """Subfeature column E should have outline level 1."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "chapter": "chapter 1",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "Transfer",
                            "subfeatures_l2": [],
                        }
                    ],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Find row with Subfeature in E (outline level should be 1)
        for row in range(1, ws.max_row + 1):
            cell_e = ws.cell(row=row, column=5)  # Column E
            if cell_e.value and "Transfer" in str(cell_e.value):
                outline_level = ws.row_dimensions[row].outline_level
                assert outline_level == 1, f"E column row {row} should have outline level 1, got {outline_level}"
                break

    def test_outline_levels_f_is_two(self, temp_dir, sample_template_xlsx):
        """Subfeature column F should have outline level 2."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "chapter": "chapter 1",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "Transfer",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2_overview": "Basic Transfer",
                                    "subfeature_l2_detail": "Data transfer details",
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Find row with Subfeature in F (outline level should be 2)
        for row in range(1, ws.max_row + 1):
            cell_f = ws.cell(row=row, column=6)  # Column F
            if cell_f.value and "Basic Transfer" in str(cell_f.value):
                outline_level = ws.row_dimensions[row].outline_level
                assert outline_level == 2, f"F column row {row} should have outline level 2, got {outline_level}"
                break

    def test_outline_levels_g_is_three(self, temp_dir, sample_template_xlsx):
        """Subfeature column G should have outline level 3."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "chapter": "chapter 1",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "Transfer",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2_overview": "Basic Transfer",
                                    "subfeature_l2_detail": "Data transfer details",
                                    "subfeatures_l3": [
                                        {
                                            "subfeature_l3": "Detailed Transfer",
                                        }
                                    ],
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Find row with Subfeature in G (outline level should be 3)
        for row in range(1, ws.max_row + 1):
            cell_g = ws.cell(row=row, column=7)  # Column G
            if cell_g.value and "Detailed Transfer" in str(cell_g.value):
                outline_level = ws.row_dimensions[row].outline_level
                assert outline_level == 3, f"G column row {row} should have outline level 3, got {outline_level}"
                break

    def test_font_is_applied(self, temp_dir, sample_template_xlsx):
        """Font should be applied to data cells."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "chapter": "chapter 1",
                    "subfeatures_l1": [],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Find row with Feature and check font
        for row in range(1, ws.max_row + 1):
            cell_d = ws.cell(row=row, column=4)
            if cell_d.value and "Data Path" in str(cell_d.value):
                font = cell_d.font
                assert font.name == "\u5fae\u8f6f\u96c5\u9ed1", f"Font should be Microsoft YaHei, got {font.name}"
                assert font.size == 10, f"Font size should be 10, got {font.size}"
                break

    def test_borders_applied_to_data_region(self, temp_dir, sample_template_xlsx):
        """Borders should be applied to the data region."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "chapter": "chapter 1",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "Transfer",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2_overview": "Basic",
                                    "subfeature_l2_detail": "Details",
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Find a data cell and check border
        for row in range(1, ws.max_row + 1):
            cell_h = ws.cell(row=row, column=8)  # Column H (description)
            if cell_h.value:
                # Check border exists (should have thin border)
                border = cell_h.border
                assert border.left is not None or border.right is not None, "Border should be applied"
                break

    def test_h_to_w_columns_at_min_granularity_with_g(self, temp_dir, sample_template_xlsx):
        """H-W columns should only appear at minimum granularity (G level)."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "chapter": "chapter 1",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "Transfer",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2_overview": "Basic",
                                    "subfeature_l2_detail": "Details",
                                    "subfeatures_l3": [
                                        {
                                            "subfeature_l3": "Detailed",
                                            "remarks": "Test remarks",
                                            "stimulus": "Test stimulus",
                                            "checking": "Test checking",
                                        }
                                    ],
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # H-W columns should only have data at rows with G column filled
        for row in range(1, ws.max_row + 1):
            cell_h = ws.cell(row=row, column=8)  # Column H
            cell_g = ws.cell(row=row, column=7)  # Column G

            # If H has data, G should also have data (minimum granularity check)
            if cell_h.value:
                assert cell_g.value is not None, f"Row {row}: H column has data but G is empty (minimum granularity violation)"

    def test_h_to_w_columns_at_min_granularity_without_g(self, temp_dir, sample_template_xlsx):
        """H-W columns should appear at F level if G does not exist."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "chapter": "chapter 1",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "Transfer",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2_overview": "Basic",
                                    "subfeature_l2_detail": "Details",
                                    "remarks": "Test remarks",
                                    "stimulus": "Test stimulus",
                                    "checking": "Test checking",
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # H-W columns should only have data at rows with F column filled (since no G)
        for row in range(1, ws.max_row + 1):
            cell_h = ws.cell(row=row, column=8)  # Column H
            cell_f = ws.cell(row=row, column=6)  # Column F

            # If H has data, F should also have data (minimum granularity check)
            if cell_h.value:
                assert cell_f.value is not None, f"Row {row}: H column has data but F is empty"

    def test_full_hierarchy_write(self, temp_dir, sample_template_xlsx, expected_ch1_json):
        """Test writing full hierarchy from expected_ch1_json structure."""
        features_data = expected_ch1_json
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Verify feature exists (Chinese: "数据通路特性")
        found_feature = False
        found_subfeature = False
        for row in range(1, ws.max_row + 1):
            cell_d = ws.cell(row=row, column=4)
            cell_e = ws.cell(row=row, column=5)
            # Check for Chinese feature name
            if cell_d.value and "\u6570\u636e" in str(cell_d.value):  # 数据
                found_feature = True
            # Check for Chinese subfeature name
            if cell_e.value and "\u4f20\u8f93" in str(cell_e.value):  # 传输
                found_subfeature = True

        assert found_feature, "Feature not found in output"
        assert found_subfeature, "Subfeature not found in output"
