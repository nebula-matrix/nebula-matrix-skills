"""Tests for validator module - Output Validator."""
import sys
from pathlib import Path

import pytest

# Add scripts to path
sys.path.insert(0, str(Path(__file__).parent.parent / "skills" / "testplan-func" / "scripts"))


class TestValidateHierarchy:
    """Tests for validate_hierarchy function."""

    def test_validate_hierarchy_exists(self):
        """validate_hierarchy function should exist."""
        from validator import validate_hierarchy
        assert callable(validate_hierarchy)

    def test_validate_hierarchy_returns_dict(self):
        """validate_hierarchy should return a dictionary."""
        from validator import validate_hierarchy
        features_data = {
            "module_name": "UPA",
            "features": []
        }
        result = validate_hierarchy(features_data)
        assert isinstance(result, dict)

    def test_validate_hierarchy_has_valid_field(self):
        """validate_hierarchy should return dict with 'valid' field."""
        from validator import validate_hierarchy
        features_data = {
            "module_name": "UPA",
            "features": []
        }
        result = validate_hierarchy(features_data)
        assert "valid" in result

    def test_validate_hierarchy_has_errors_field(self):
        """validate_hierarchy should return dict with 'errors' field."""
        from validator import validate_hierarchy
        features_data = {
            "module_name": "UPA",
            "features": []
        }
        result = validate_hierarchy(features_data)
        assert "errors" in result
        assert isinstance(result["errors"], list)

    def test_validate_hierarchy_passes_valid_data(self):
        """validate_hierarchy should pass valid hierarchy data."""
        from validator import validate_hierarchy
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Spec",
            "chapter": "chapter 1",
            "features": [
                {
                    "feature": "Data Path",
                    "feature_id": "PA.001",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "Transfer",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2_overview": "Basic",
                                    "subfeature_l2_detail": "Details",
                                }
                            ]
                        }
                    ]
                }
            ]
        }
        result = validate_hierarchy(features_data)
        assert result["valid"] is True

    def test_validate_hierarchy_catches_missing_feature(self):
        """validate_hierarchy should catch missing feature field."""
        from validator import validate_hierarchy
        features_data = {
            "module_name": "UPA",
            "features": [
                {
                    "feature_id": "PA.001",
                    "subfeatures_l1": []
                }
            ]
        }
        result = validate_hierarchy(features_data)
        assert result["valid"] is False
        assert len(result["errors"]) > 0


class TestValidateOutlineLevels:
    """Tests for validate_outline_levels function."""

    def test_validate_outline_levels_exists(self):
        """validate_outline_levels function should exist."""
        from validator import validate_outline_levels
        assert callable(validate_outline_levels)

    def test_validate_outline_levels_returns_dict(self, temp_dir, sample_template_xlsx):
        """validate_outline_levels should return a dictionary."""
        from validator import validate_outline_levels
        from openpyxl import load_workbook
        import shutil

        # Copy template to temp location
        output_path = temp_dir / "test.xlsx"
        shutil.copy(sample_template_xlsx, output_path)

        result = validate_outline_levels(output_path)
        assert isinstance(result, dict)

    def test_validate_outline_levels_has_valid_field(self, temp_dir, sample_template_xlsx):
        """validate_outline_levels should return dict with 'valid' field."""
        from validator import validate_outline_levels
        import shutil

        output_path = temp_dir / "test.xlsx"
        shutil.copy(sample_template_xlsx, output_path)

        result = validate_outline_levels(output_path)
        assert "valid" in result

    def test_validate_outline_levels_validates_range(self, temp_dir, sample_template_xlsx):
        """validate_outline_levels should check outline levels are 0-3."""
        from validator import validate_outline_levels
        import shutil

        output_path = temp_dir / "test.xlsx"
        shutil.copy(sample_template_xlsx, output_path)

        result = validate_outline_levels(output_path)
        # Should pass for valid template
        assert "errors" in result


class TestValidateHWColumns:
    """Tests for validate_hw_columns function."""

    def test_validate_hw_columns_exists(self):
        """validate_hw_columns function should exist."""
        from validator import validate_hw_columns
        assert callable(validate_hw_columns)

    def test_validate_hw_columns_returns_dict(self, temp_dir, sample_template_xlsx):
        """validate_hw_columns should return a dictionary."""
        from validator import validate_hw_columns
        import shutil

        output_path = temp_dir / "test.xlsx"
        shutil.copy(sample_template_xlsx, output_path)

        result = validate_hw_columns(output_path)
        assert isinstance(result, dict)

    def test_validate_hw_columns_has_valid_field(self, temp_dir, sample_template_xlsx):
        """validate_hw_columns should return dict with 'valid' field."""
        from validator import validate_hw_columns
        import shutil

        output_path = temp_dir / "test.xlsx"
        shutil.copy(sample_template_xlsx, output_path)

        result = validate_hw_columns(output_path)
        assert "valid" in result


class TestValidateXlsxOutput:
    """Tests for validate_xlsx_output function."""

    def test_validate_xlsx_output_exists(self):
        """validate_xlsx_output function should exist."""
        from validator import validate_xlsx_output
        assert callable(validate_xlsx_output)

    def test_validate_xlsx_output_returns_dict(self, temp_dir, sample_template_xlsx):
        """validate_xlsx_output should return a dictionary."""
        from validator import validate_xlsx_output
        import shutil

        output_path = temp_dir / "test.xlsx"
        shutil.copy(sample_template_xlsx, output_path)

        result = validate_xlsx_output(output_path)
        assert isinstance(result, dict)

    def test_validate_xlsx_output_combines_checks(self, temp_dir, sample_template_xlsx):
        """validate_xlsx_output should combine all validation checks."""
        from validator import validate_xlsx_output
        import shutil

        output_path = temp_dir / "test.xlsx"
        shutil.copy(sample_template_xlsx, output_path)

        result = validate_xlsx_output(output_path)
        # Should have combined results
        assert "valid" in result
        assert "errors" in result

    def test_validate_xlsx_output_raises_for_missing_file(self, temp_dir):
        """validate_xlsx_output should raise for missing file."""
        from validator import validate_xlsx_output

        with pytest.raises(FileNotFoundError):
            validate_xlsx_output(temp_dir / "nonexistent.xlsx")


class TestValidateMinimumGranularity:
    """Tests for minimum granularity validation."""

    def test_hw_only_at_g_when_g_exists(self, temp_dir, sample_template_xlsx):
        """H-W columns should only exist at G level when G is present."""
        from validator import validate_hw_columns
        from openpyxl import Workbook
        from openpyxl.styles import Font

        # Create a workbook with G level having H-W data
        wb = Workbook()
        ws = wb.active

        # Row with D column (level 0)
        ws.cell(row=1, column=4, value="Feature")
        ws.row_dimensions[1].outline_level = 0

        # Row with E column (level 1)
        ws.cell(row=2, column=5, value="Subfeature L1")
        ws.row_dimensions[2].outline_level = 1

        # Row with F column (level 2)
        ws.cell(row=3, column=6, value="Subfeature L2")
        ws.row_dimensions[3].outline_level = 2

        # Row with G column (level 3) - should have H-W
        ws.cell(row=4, column=7, value="Subfeature L3")
        ws.cell(row=4, column=8, value="Remarks")
        ws.row_dimensions[4].outline_level = 3

        output_path = temp_dir / "test_granularity.xlsx"
        wb.save(output_path)
        wb.close()

        result = validate_hw_columns(output_path)
        # Row 4 has H at G level - should be valid
        assert result["valid"] is True

    def test_hw_at_f_when_no_g(self, temp_dir, sample_template_xlsx):
        """H-W columns should be at F level when G does not exist."""
        from validator import validate_hw_columns
        from openpyxl import Workbook

        # Create a workbook without G level, H-W at F
        wb = Workbook()
        ws = wb.active

        # D level
        ws.cell(row=1, column=4, value="Feature")
        ws.row_dimensions[1].outline_level = 0

        # E level
        ws.cell(row=2, column=5, value="Subfeature L1")
        ws.row_dimensions[2].outline_level = 1

        # F level - minimum granularity
        ws.cell(row=3, column=6, value="Subfeature L2")
        ws.cell(row=3, column=8, value="Remarks")  # H column
        ws.row_dimensions[3].outline_level = 2

        output_path = temp_dir / "test_f_granularity.xlsx"
        wb.save(output_path)
        wb.close()

        result = validate_hw_columns(output_path)
        # H at F level with no G - should be valid
        assert result["valid"] is True


class TestValidateFeatureStructure:
    """Tests for feature structure validation."""

    def test_feature_must_have_feature_or_feature_id(self):
        """Feature must have at least feature or feature_id field."""
        from validator import validate_hierarchy

        features_data = {
            "module_name": "UPA",
            "features": [
                {"subfeatures_l1": []}  # Missing both feature and feature_id
            ]
        }
        result = validate_hierarchy(features_data)
        assert result["valid"] is False

    def test_subfeature_l1_must_have_name(self):
        """L1 subfeature must have subfeature_l1 field."""
        from validator import validate_hierarchy

        features_data = {
            "module_name": "UPA",
            "features": [
                {
                    "feature": "Test",
                    "subfeatures_l1": [
                        {"subfeatures_l2": []}  # Missing subfeature_l1
                    ]
                }
            ]
        }
        result = validate_hierarchy(features_data)
        assert result["valid"] is False

    def test_subfeature_l2_must_have_identifier(self):
        """L2 subfeature must have at least one identifier field."""
        from validator import validate_hierarchy

        features_data = {
            "module_name": "UPA",
            "features": [
                {
                    "feature": "Test",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "L1",
                            "subfeatures_l2": [
                                {}  # Missing all identifiers
                            ]
                        }
                    ]
                }
            ]
        }
        result = validate_hierarchy(features_data)
        assert result["valid"] is False
