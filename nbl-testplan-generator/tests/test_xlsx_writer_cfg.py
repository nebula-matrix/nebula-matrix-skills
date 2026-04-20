"""Tests for testplan-cfg xlsx_writer module."""
import importlib.util
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Explicitly load xlsx_writer from testplan-cfg/scripts
_cfg_xlsx_writer = Path(__file__).parent.parent / "skills" / "testplan-cfg" / "scripts" / "xlsx_writer.py"
_spec = importlib.util.spec_from_file_location("xlsx_writer_cfg", _cfg_xlsx_writer)
xlsx_writer_cfg = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(xlsx_writer_cfg)

write_ch1_xlsx = xlsx_writer_cfg.write_ch1_xlsx


class TestClearTemplateData:
    """Tests for clearing template data before writing."""

    def test_template_data_cleared_from_row_6(self, temp_dir, sample_template_xlsx):
        """Template example data should be cleared starting from row 6."""
        features_data = {
            "module_name": "UPA",
            "spec_name": "UPA Module Spec",
            "chapter": "chapter 2 配置特性",
            "features": [
                {
                    "feature": "upa_ctrl",
                    "feature_id": "",
                    "chapter": "chapter 2",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "en",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2_overview": "bit[0] 使能控制",
                                    "remarks": "",
                                    "stimulus": "【配置】配置 upa_ctrl.en",
                                    "checking": "随机用例",
                                    "path": "Group:$unit::upa_fcov::upa_ctrl_cg.en_cp",
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output_path = temp_dir / "output.xlsx"
        write_ch1_xlsx(features_data, sample_template_xlsx, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Template example data markers that should NOT exist in output
        # These are from rows 6-10 of the template
        template_markers = [
            "chapter x.x",  # Row 6 Col B
            "Feature0",     # Row 7 Col E
            "feature0_0",   # Row 8 Col F
            "feature0_1",   # Row 9 Col F
            "feature0_2",   # Row 10 Col F
        ]

        # Check all cells in rows 6-10 for template marker strings
        for row in range(6, 11):
            for col in range(1, 24):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    cell_str = str(cell_value)
                    for marker in template_markers:
                        assert marker not in cell_str, (
                            f"Template example data '{marker}' found at row {row}, col {col}. "
                            f"Template data should be cleared before writing new data."
                        )

        wb.close()
