"""Tests for combine_writer module."""
from pathlib import Path

import pytest

from writers.combine_writer import write_testplan_xlsx


class TestWriteTestplanXlsx:
    def test_creates_file(self, temp_dir, sample_template_xlsx):
        """Should create output xlsx file."""
        data = {
            "module_name": "UPA",
            "spec_name": "UPA_spec",
            "chapter": "chapter 1 功能特性",
            "features": [],
        }
        output = temp_dir / "test_output.xlsx"
        result = write_testplan_xlsx(data, sample_template_xlsx, output)
        assert output.exists()
        assert result == output.resolve()

    def test_writes_feature_in_column_d(self, temp_dir, sample_template_xlsx):
        """Feature name should appear in column D."""
        data = {
            "module_name": "UPA",
            "spec_name": "UPA_spec",
            "chapter": "chapter 1 功能特性",
            "features": [
                {
                    "feature": "报文编辑范围",
                    "feature_id": "PA.001",
                    "subfeatures_l1": [],
                }
            ],
        }
        output = temp_dir / "test_output.xlsx"
        write_testplan_xlsx(data, sample_template_xlsx, output)
        import openpyxl
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        found = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=4).value
            if val and "报文编辑范围" in str(val):
                found = True
                break
        assert found, "Feature not found in column D"

    def test_writes_subfeature_l1_in_column_e(self, temp_dir, sample_template_xlsx):
        data = {
            "module_name": "UPA",
            "spec_name": "UPA_spec",
            "chapter": "chapter 1 功能特性",
            "features": [
                {
                    "feature": "报文编辑范围",
                    "feature_id": "PA.001",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "替换动作",
                            "subfeatures_l2": [],
                        }
                    ],
                }
            ],
        }
        output = temp_dir / "test_output.xlsx"
        write_testplan_xlsx(data, sample_template_xlsx, output)
        import openpyxl
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        found = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=5).value
            if val and "替换动作" in str(val):
                found = True
                break
        assert found, "Subfeature L1 not found in column E"

    def test_data_row_has_borders(self, temp_dir, sample_template_xlsx):
        data = {
            "module_name": "UPA",
            "spec_name": "UPA_spec",
            "chapter": "chapter 1 功能特性",
            "features": [
                {
                    "feature": "F",
                    "feature_id": "PA.001",
                    "_eng_id": "test_feat",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "E",
                            "_eng_id": "test_sub",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2": "F2overview",
                                    "_confidence": "confirmed",
                                    "subfeatures_l3": [
                                        {
                                            "subfeature_l3": "G3detail",
                                            "remarks": "test",
                                            "stimulus": "【配置】test\n【激励】test",
                                            "checking": "by_checker",
                                            "coverage": "test",
                                            "priority": "HIGH",
                                            "activity": "BT",
                                            "path": "",
                                        }
                                    ],
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output = temp_dir / "test_output.xlsx"
        write_testplan_xlsx(data, sample_template_xlsx, output)
        import openpyxl
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=8).value:
                border = ws.cell(row=row, column=4).border
                assert border is not None
                assert border.left.style is not None
                break
        else:
            pytest.fail("No data row found")

    def test_eng_id_path(self, temp_dir, sample_template_xlsx):
        """_eng_id should be used to build W column path."""
        data = {
            "module_name": "upa",
            "spec_name": "UPA_spec",
            "chapter": "chapter 1 功能特性",
            "features": [
                {
                    "feature": "报文编辑",
                    "feature_id": "PA.001",
                    "_eng_id": "pkt_edt",
                    "subfeatures_l1": [
                        {
                            "subfeature_l1": "替换",
                            "_eng_id": "repl",
                            "subfeatures_l2": [
                                {
                                    "subfeature_l2": "替换测试",
                                    "_confidence": "confirmed",
                                    "subfeatures_l3": [
                                        {
                                            "subfeature_l3": "正常替换",
                                            "remarks": "测试",
                                            "stimulus": "测试",
                                            "checking": "by_checker",
                                            "coverage": "测试",
                                            "priority": "HIGH",
                                            "activity": "BT",
                                            "path": "",
                                        }
                                    ],
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        output = temp_dir / "test_output.xlsx"
        write_testplan_xlsx(data, sample_template_xlsx, output)
        import openpyxl
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=23).value
            if val and "cg_pkt_edt" in str(val):
                assert "cp_repl" in str(val)
                return
        pytest.fail("W column path not using _eng_id correctly")
