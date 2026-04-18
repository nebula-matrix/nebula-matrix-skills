#!/usr/bin/env python3
"""cfg_writer.py - Configuration feature generation and appending.

This module handles Ch2 (配置特性) generation:
1. Extract covered registers from existing func xlsx
2. Filter out excluded registers (*_int_*, *_fifo_*, etc.)
3. Group similar registers (_0/_1/_2 -> _0/1/2)
4. Cross-reference with covered registers
5. Append Ch2 to func xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl


# Exclusion patterns for Ch2 registers
CFG_EXCLUDE_PATTERNS: list[str] = [
    r".*_int_status$",
    r".*_int_mask$",
    r".*_int_set$",
    r".*_int_.*$",
    r".*_car_ctrl$",
    r".*_fifo_th$",
    r".*_fifo_.*$",
    r".*_spare\d*$",
    r".*_cnt$",
    r".*_fsm\d*$",
    r".*_err_info$",
    r".*_bp_state$",
    r".*_bp_history$",
    r".*_test$",
    r".*_init_done$",
    r".*_init_start$",
]

_CFG_EXCLUDE_COMPILED = [re.compile(p) for p in CFG_EXCLUDE_PATTERNS]

# Valid access types for configuration registers
VALID_CFG_ACCESS = {"RW", "RWC", "WO", "RWW"}


def _is_excluded_register(reg_name: str) -> bool:
    """Check if register should be excluded from Ch2."""
    for pat in _CFG_EXCLUDE_COMPILED:
        if pat.match(reg_name):
            return True
    return False


def extract_covered_regs_from_xlsx(func_path: str) -> set[str]:
    """Extract covered register names from existing func xlsx.

    Scans the stimulus and coverage columns for register references.

    Parameters
    ----------
    func_path : str
        Path to the func xlsx file.

    Returns
    -------
    set[str]
        Set of register names found in the xlsx.
    """
    wb = openpyxl.load_workbook(func_path, data_only=True)
    ws = wb.active

    covered: set[str] = set()

    # Find header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for j, cell in enumerate(row):
            if cell and "stimulus" in str(cell).lower():
                header_row = i
                break
        if header_row:
            break

    if header_row is None:
        wb.close()
        return covered

    # Find column indices
    headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    col_map = {}
    for j, h in enumerate(headers):
        if h:
            h_lower = str(h).lower()
            if "stimulus" in h_lower:
                col_map["stimulus"] = j + 1
            elif "coverage" in h_lower:
                col_map["coverage"] = j + 1

    # Scan data rows for register names
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        stimulus = str(row[col_map.get("stimulus", 5) - 1] or "")
        coverage = str(row[col_map.get("coverage", 7) - 1] or "")
        text = stimulus + " " + coverage

        # Find register names (upa_xxx pattern)
        for match in re.finditer(r"(upa_\w+)", text):
            reg_name = match.group(1)
            # Normalize: strip trailing _0/_1 if present
            base_match = re.match(r"^(.+?)_\d+$", reg_name)
            if base_match:
                covered.add(base_match.group(1))
            covered.add(reg_name)

    wb.close()
    return covered


def filter_cfg_registers(registers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Filter registers for Ch2 configuration features.

    - Excludes registers matching exclusion patterns (*_int_*, *_fifo_*, etc.)
    - Keeps only fields with valid access types (RW, RWC, WO, RWW)

    Parameters
    ----------
    registers : list[dict]
        List of register dictionaries from reg_to_json output.

    Returns
    -------
    list[dict]
        Filtered registers with valid configuration fields.
    """
    result: list[dict[str, Any]] = []

    for reg in registers:
        reg_name = reg.get("name", "")

        # Skip excluded registers
        if _is_excluded_register(reg_name):
            continue

        # Filter fields to only valid access types
        valid_fields = []
        for field in reg.get("fields", []):
            access = field.get("access", "").upper()
            if access in VALID_CFG_ACCESS:
                valid_fields.append(field)

        # Only include register if it has valid fields
        if valid_fields:
            result.append({
                **reg,
                "fields": valid_fields,
            })

    return result


def group_similar_registers(registers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Group registers that differ only by a trailing numeric suffix.

    E.g. upa_stage_0, upa_stage_1, upa_stage_2 -> upa_stage_0/1/2

    Parameters
    ----------
    registers : list[dict]
        List of register dictionaries.

    Returns
    -------
    list[dict]
        Grouped registers with merged names.
    """
    groups: dict[str, list[dict[str, Any]]] = {}
    order: list[str] = []

    for reg in registers:
        name = reg["name"]
        m = re.match(r"^(.+?)_(\d+)$", name)
        if m:
            base = m.group(1)
        else:
            base = name

        if base not in groups:
            groups[base] = []
            order.append(base)
        groups[base].append(reg)

    result: list[dict[str, Any]] = []
    for base in order:
        members = groups[base]
        if len(members) == 1:
            result.append(members[0])
        else:
            suffixes = []
            for m_reg in members:
                m_match = re.match(r"^.+?_(\d+)$", m_reg["name"])
                suffixes.append(m_match.group(1) if m_match else "")

            merged_name = f"{base}_{'/'.join(suffixes)}"

            # Merge fields from all members
            all_fields: list[dict[str, Any]] = []
            seen: set[str] = set()
            for m_reg in members:
                for field in m_reg.get("fields", []):
                    key = field.get("name", "")
                    if key and key not in seen:
                        seen.add(key)
                        all_fields.append(field)

            result.append({
                "name": merged_name,
                "offset": members[0].get("offset", ""),
                "fields": all_fields,
                "description": members[0].get("description", ""),
            })

    return result


def cross_ref_cfg(
    ch2_regs: list[dict[str, Any]],
    covered_regs: set[str],
) -> list[dict[str, Any]]:
    """Cross-reference Ch2 registers with covered registers.

    Parameters
    ----------
    ch2_regs : list[dict]
        Ch2 register list (after filtering and grouping).
    covered_regs : set[str]
        Set of already covered register names.

    Returns
    -------
    list[dict]
        Cross-referenced registers with skip flag and stimulus.
    """
    result: list[dict[str, Any]] = []

    for reg in ch2_regs:
        reg_name = reg["name"]
        fields = reg.get("fields", [])

        # Check if this register (or its base) is covered
        is_covered = reg_name in covered_regs
        if not is_covered:
            # Check if grouped name matches
            base_match = re.match(r"^(.+?)_\d+(/\d+)*$", reg_name)
            if base_match:
                base = base_match.group(1)
                if base in covered_regs:
                    is_covered = True
                # Also check individual members
                for covered in covered_regs:
                    if covered.startswith(base + "_"):
                        is_covered = True
                        break

        if is_covered:
            # Mark as skip
            result.append({
                "name": reg_name,
                "skip": True,
                "stimulus": "",
                "checking": "",
                "coverage": "",
                "path": "",
                "remarks": "已在功能特性中覆盖",
            })
        else:
            # Generate stimulus
            desc_parts = []
            config_parts = []
            for field in fields:
                fname = field.get("name", "")
                fdesc = field.get("description", "")
                frange = field.get("range", "")
                faccess = field.get("access", "")

                if fname and fname.lower() != "rsv":
                    if fdesc:
                        desc_parts.append(f"{frange} {fname}: {fdesc}")

                    # Generate config for RW/RWC/WO fields
                    if faccess.upper() in VALID_CFG_ACCESS:
                        config_parts.append(f"配置 {reg_name}.{fname} ({frange})")

            stimulus = "【配置】" + "；".join(config_parts) if config_parts else ""

            result.append({
                "name": reg_name,
                "skip": False,
                "stimulus": stimulus,
                "checking": "",
                "coverage": "",
                "path": "",
                "remarks": "; ".join(desc_parts) if desc_parts else "",
            })

    return result


def append_cfg_xlsx(
    func_path: str,
    reg_data: dict[str, Any],
    output_path: str,
) -> dict[str, int]:
    """Append Ch2 configuration features to func xlsx.

    Parameters
    ----------
    func_path : str
        Path to the existing func xlsx.
    reg_data : dict
        Register JSON data from reg_to_json output.
    output_path : str
        Path for the output xlsx.

    Returns
    -------
    dict[str, int]
        Statistics: total_registers, covered_registers, new_registers.
    """
    # Load func xlsx
    wb = openpyxl.load_workbook(func_path)
    ws = wb.active

    # Extract covered registers
    covered_regs = extract_covered_regs_from_xlsx(func_path)

    # Process registers
    registers = reg_data.get("registers", [])
    filtered = filter_cfg_registers(registers)
    grouped = group_similar_registers(filtered)
    xrefed = cross_ref_cfg(grouped, covered_regs)

    # Find the last data row
    last_row = ws.max_row
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=4).value:  # Column D has feature names
            last_row = row
            break

    # Append Ch2 data
    next_row = last_row + 1

    # Add Ch2 header row
    ws.cell(row=next_row, column=4, value="chapter 2 配置特性")
    ws.row_dimensions[next_row].outline_level = 0
    next_row += 1

    covered_count = 0
    new_count = 0

    for reg_info in xrefed:
        reg_name = reg_info["name"]
        is_skip = reg_info.get("skip", False)

        # Register row
        ws.cell(row=next_row, column=4, value=reg_name)
        ws.row_dimensions[next_row].outline_level = 1
        next_row += 1

        # Subfeature row with stimulus
        ws.cell(row=next_row, column=6, value="")  # F column: overview
        ws.cell(row=next_row, column=9, value=reg_info.get("stimulus", ""))  # I column: stimulus
        ws.cell(row=next_row, column=10, value=reg_info.get("checking", ""))  # J column: checking
        ws.cell(row=next_row, column=11, value=reg_info.get("coverage", ""))  # K column: coverage
        ws.cell(row=next_row, column=12, value="HIGH" if not is_skip else "")  # L column: priority
        ws.cell(row=next_row, column=13, value="BT")  # M column: activity
        ws.cell(row=next_row, column=23, value=reg_info.get("path", ""))  # W column: path

        if is_skip:
            covered_count += 1
        else:
            new_count += 1

        ws.row_dimensions[next_row].outline_level = 2
        next_row += 1

    wb.save(output_path)
    wb.close()

    return {
        "total_registers": len(xrefed),
        "covered_registers": covered_count,
        "new_registers": new_count,
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate and append Ch2 configuration features")
    parser.add_argument("--func", required=True, help="Path to existing func xlsx")
    parser.add_argument("--reg", required=True, help="Path to register JSON")
    parser.add_argument("--output", required=True, help="Output xlsx path")
    args = parser.parse_args()

    with open(args.reg, "r", encoding="utf-8") as f:
        reg_data = json.load(f)

    stats = append_cfg_xlsx(args.func, reg_data, args.output)
    print(f"Done: {stats}")
