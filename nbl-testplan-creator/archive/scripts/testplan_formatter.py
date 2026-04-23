#!/usr/bin/env python3
"""
测试计划格式化脚本
从统一的 features.json 生成标准格式的测试计划文档
支持三级Feature结构：Feature -> Sub-Feature -> Testpoint
"""

import json
import argparse
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime


class TestPlanFormatter:
    """测试计划格式化器"""

    def __init__(self):
        pass

    def load_features_json(self, json_path: str) -> Dict[str, Any]:
        """从 features.json 加载完整的三级结构"""
        path = Path(json_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {path}")

        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def generate_markdown_tree(self, data: Dict[str, Any]) -> str:
        """生成 Markdown 树状格式的测试计划

        格式：
        ## Feature名称

        - description: xxx
        - priority: HIGH
        - sections_covered: S001,S002

        ### Sub-Feature名称

        - sub_feature_name: xxx
        - description: xxx
        - spec_id: 【UVN.001.001】
        - priority: HIGH

        | tp_name | source | stimulus | checking | coverage_requirements | priority | category | path2source |
        |---------|--------|----------|----------|-----------------------|----------|----------|-------------|
        | xxx | xxx | xxx | xxx | xxx | xxx | xxx | xxx |
        """
        lines = []
        document_title = data.get('document_title') or 'IC模块验证测试计划'

        # 标题与文档级元数据（非 heading，用列表行展示）
        features = data.get('features', [])
        total_sub_features = sum(len(f.get('sub_features', [])) for f in features)
        total_testpoints = sum(
            len(sf.get('testpoints', []))
            for f in features
            for sf in f.get('sub_features', [])
        )

        lines.append(f"# {document_title}")
        lines.append("")
        lines.append(f"- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"- 总Feature数: {len(features)} | 总Sub-Feature数: {total_sub_features} | 总测试点数: {total_testpoints}")
        lines.append("")

        # 按 Feature 组织内容
        for fi, feature in enumerate(features, 1):
            feature_name = feature.get('feature_name', '')
            description = feature.get('description', '')
            priority = feature.get('priority', '')
            sections_covered = feature.get('sections_covered', [])

            # Feature 标题（带编号）
            lines.append(f"## {fi}. {feature_name}")
            lines.append("")

            # Feature 元数据
            if description:
                lines.append(f"- description: {description}")
            if priority:
                lines.append(f"- priority: {priority}")
            if sections_covered:
                lines.append(f"- sections_covered: {','.join(sections_covered)}")
            lines.append("")

            # Sub-Features
            for sfi, sub_feature in enumerate(feature.get('sub_features', []), 1):
                sub_feature_name = sub_feature.get('sub_feature_name', '')
                sub_desc = sub_feature.get('description', '')
                spec_id = sub_feature.get('spec_id', '')
                sub_priority = sub_feature.get('priority', '')
                testpoints = sub_feature.get('testpoints', [])

                # Sub-Feature 标题（带编号，使用 sub_feature_name 或 description 作为标题）
                title = sub_feature_name or sub_desc
                lines.append(f"### {fi}.{sfi} {title}")
                lines.append("")

                # Sub-Feature 元数据
                if sub_feature_name:
                    lines.append(f"- sub_feature_name: {sub_feature_name}")
                if sub_desc:
                    lines.append(f"- description: {sub_desc}")
                if spec_id:
                    lines.append(f"- spec_id: {spec_id}")
                if sub_priority:
                    lines.append(f"- priority: {sub_priority}")

                # 元数据和表格之间必须有空行
                if testpoints:
                    lines.append("")
                    lines.append("| tp_name | source | stimulus | checking | coverage_requirements | priority | category | path2source |")
                    lines.append("|---------|--------|----------|----------|-----------------------|----------|----------|-------------|")

                    for tp in testpoints:
                        tp_name = tp.get('tp_name', '')
                        source = tp.get('source', '')
                        stimulus = tp.get('stimulus', '').replace('\n', '<br/>')
                        checking = tp.get('checking', '')
                        coverage = tp.get('coverage_requirements', '')
                        tp_priority = tp.get('priority', '')
                        category = tp.get('category', '')
                        path2source = tp.get('path2source', '')
                        skip = tp.get('skip', False)

                        # skip 标记
                        if skip:
                            tp_name = f'[SKIP] {tp_name}'

                        # 转义管道符防止破坏表格
                        tp_name_safe = str(tp_name).replace('|', '\\|')
                        source_safe = str(source).replace('|', '\\|')
                        stimulus_safe = str(stimulus).replace('|', '\\|')
                        coverage_safe = str(coverage).replace('|', '\\|')
                        path2source_safe = str(path2source).replace('|', '\\|')

                        lines.append(f"| {tp_name_safe} | {source_safe} | {stimulus_safe} | {checking} | {coverage_safe} | {tp_priority} | {category} | {path2source_safe} |")

                # Sub-Feature 之间空行分开
                lines.append("")

        return '\n'.join(lines)

    def generate_excel(self, data: Dict[str, Any], output_path: str):
        """生成Excel格式的测试计划"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("错误：需要安装 openpyxl 库")
            print("请运行: pip install openpyxl")
            return False

        wb = Workbook()
        ws = wb.active
        ws.title = "测试计划"

        # 样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # 表头
        headers = ["Feature名称", "Sub-Feature描述", "规格编号", "Testpoint名称",
                   "来源", "激励", "检查方式", "覆盖率需求", "优先级", "类别", "路径"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        row = 2
        for feature in data.get('features', []):
            feature_name = feature.get('feature_name', '')

            for sub_feature in feature.get('sub_features', []):
                sub_desc = sub_feature.get('description', '')
                spec_id = sub_feature.get('spec_id', '')

                for tp in sub_feature.get('testpoints', []):
                    tp_name = tp.get('tp_name', '')
                    source = tp.get('source', '')
                    stimulus = tp.get('stimulus', '')
                    checking = tp.get('checking', '')
                    coverage = tp.get('coverage_requirements', '')
                    priority = tp.get('priority', '')
                    category = tp.get('category', '')
                    path = tp.get('path2source', '')

                    ws.cell(row=row, column=1, value=feature_name)
                    ws.cell(row=row, column=2, value=sub_desc)
                    ws.cell(row=row, column=3, value=spec_id)
                    ws.cell(row=row, column=4, value=tp_name)
                    ws.cell(row=row, column=5, value=source)
                    ws.cell(row=row, column=6, value=stimulus)
                    ws.cell(row=row, column=7, value=checking)
                    ws.cell(row=row, column=8, value=coverage)
                    ws.cell(row=row, column=9, value=priority)
                    ws.cell(row=row, column=10, value=category)
                    ws.cell(row=row, column=11, value=path)

                    row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 30

        wb.save(output_path)
        return True

    def generate_csv(self, data: Dict[str, Any]) -> str:
        """生成CSV格式的测试计划"""
        lines = []
        lines.append("Feature名称,Sub-Feature描述,规格编号,Testpoint名称,来源,激励,检查方式,覆盖率需求,优先级,类别,路径")

        for feature in data.get('features', []):
            feature_name = feature.get('feature_name', '')

            for sub_feature in feature.get('sub_features', []):
                sub_desc = sub_feature.get('description', '')
                spec_id = sub_feature.get('spec_id', '')

                for tp in sub_feature.get('testpoints', []):
                    tp_name = tp.get('tp_name', '')
                    source = tp.get('source', '')
                    stimulus = tp.get('stimulus', '').replace('\n', ' ')
                    checking = tp.get('checking', '')
                    coverage = tp.get('coverage_requirements', '')
                    priority = tp.get('priority', '')
                    category = tp.get('category', '')
                    path = tp.get('path2source', '')

                    # CSV 转义：处理逗号和引号
                    def csv_escape(s):
                        s = str(s)
                        if ',' in s or '"' in s or '\n' in s:
                            s = s.replace('"', '""')
                            return f'"{s}"'
                        return s

                    line = f"{csv_escape(feature_name)},{csv_escape(sub_desc)},{csv_escape(spec_id)},{csv_escape(tp_name)},{csv_escape(source)},{csv_escape(stimulus)},{csv_escape(checking)},{csv_escape(coverage)},{csv_escape(priority)},{csv_escape(category)},{csv_escape(path)}"
                    lines.append(line)

        return '\n'.join(lines)

    def save_to_file(self, content: str, output_path: str):
        """保存内容到文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(content)


def main():
    parser = argparse.ArgumentParser(
        description='从统一的 features.json 生成测试计划',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # Markdown格式
  python3 scripts/testplan_formatter.py features.json --format markdown -o testplan.md

  # Excel格式
  python3 scripts/testplan_formatter.py features.json --format excel -o testplan.xlsx

  # CSV格式
  python3 scripts/testplan_formatter.py features.json --format csv -o testplan.csv
        """
    )

    parser.add_argument(
        'features_json',
        help='统一的 features.json 文件路径'
    )

    parser.add_argument(
        '--format',
        choices=['markdown', 'csv', 'excel'],
        default='markdown',
        help='输出格式（默认markdown）'
    )

    parser.add_argument(
        '-o', '--output',
        help='输出文件路径（可选）'
    )

    args = parser.parse_args()

    # 加载 features.json
    formatter = TestPlanFormatter()
    try:
        data = formatter.load_features_json(args.features_json)
    except FileNotFoundError as e:
        print(f"错误: {e}")
        return

    features = data.get('features', [])
    total_sub_features = sum(len(f.get('sub_features', [])) for f in features)
    total_testpoints = sum(
        len(sf.get('testpoints', []))
        for f in features
        for sf in f.get('sub_features', [])
    )
    print(f"📊 加载了 {len(features)} 个Feature，{total_sub_features} 个Sub-Feature，{total_testpoints} 个测试点")

    if args.format == 'markdown':
        content = formatter.generate_markdown_tree(data)
        if args.output:
            formatter.save_to_file(content, args.output)
            print(f"✅ Markdown测试计划已保存到: {args.output}")
        else:
            print(content)

    elif args.format == 'csv':
        content = formatter.generate_csv(data)
        if args.output:
            formatter.save_to_file(content, args.output)
            print(f"✅ CSV测试计划已保存到: {args.output}")
        else:
            print(content)

    elif args.format == 'excel':
        if not args.output:
            args.output = 'testplan.xlsx'
        if formatter.generate_excel(data, args.output):
            print(f"✅ Excel测试计划已保存到: {args.output}")


if __name__ == '__main__':
    main()
